# RP Spray Analytics
# Copyright © 2026 Ryan Phillips
# All rights reserved.
# Unauthorized copying, distribution, or resale prohibited.

import streamlit as st
st.cache_data.clear()

# -------------------------------------------------
# Per-run nonce (prevents accidental duplicate UI renders in loops)
# -------------------------------------------------
if "_rp_run_nonce" not in st.session_state:
    st.session_state["_rp_run_nonce"] = 0
st.session_state["_rp_run_nonce"] += 1
_RP_RUN_NONCE = st.session_state["_rp_run_nonce"]

import os
import json
import base64
import re
import hashlib
import httpx
import time  # anti-stuck processing lock + failsafe unlock
from datetime import datetime
import uuid
import traceback

def _write_table_two_blocks(ws, start_row, cols, row_values, split_at=None, gap=2):
    """Write a header + rows into two side-by-side blocks for landscape printing.
    - cols: list of column names
    - row_values: list of lists (each list aligns to cols)
    - split_at: index to split columns. If None, split roughly in half.
    """
    if not cols:
        return start_row

    if split_at is None:
        split_at = max(1, (len(cols) + 1) // 2)

    left_cols = cols[:split_at]
    right_cols = cols[split_at:]

    left_start_col = 1
    right_start_col = 1 + len(left_cols) + gap

    # Header
    for j, c in enumerate(left_cols, start=0):
        ws.cell(row=start_row, column=left_start_col + j, value=c)
    for j, c in enumerate(right_cols, start=0):
        ws.cell(row=start_row, column=right_start_col + j, value=c)

    # Rows
    r = start_row + 1
    for vals in row_values:
        left_vals = vals[:split_at]
        right_vals = vals[split_at:]
        for j, v in enumerate(left_vals, start=0):
            ws.cell(row=r, column=left_start_col + j, value=v)
        for j, v in enumerate(right_vals, start=0):
            ws.cell(row=r, column=right_start_col + j, value=v)
        r += 1

    return r

from typing import Optional, Tuple

import pandas as pd
from io import BytesIO

from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, CellIsRule
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from supabase import create_client, Client

SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_SERVICE_ROLE_KEY"]
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)


def hash_access_code(code: str) -> str:
    pepper = st.secrets["ACCESS_CODE_PEPPER"]
    c = (code or "").strip().upper()          # <-- THE FIX (case-insensitive)
    raw = (pepper + "|" + c).encode("utf-8")  # stable + clear separator
    return hashlib.sha256(raw).hexdigest()

def admin_set_access_code(team_slug: str, team_code: str, new_code: str) -> bool:
    team_slug = (team_slug or "").strip()
    team_code = (team_code or "").strip().upper()
    if not team_slug and not team_code:
        return False

    new_hash = hash_access_code(new_code)

    q = supabase.table("team_access").update({"code_hash": new_hash})

    # Update by whichever identifier exists (and both if both exist)
    if team_slug:
        q = q.eq("team_slug", team_slug)
    if team_code:
        q = q.eq("team_code", team_code)

    res = q.execute()
    return bool(getattr(res, "data", None))


# -----------------------------
# PATHS / FOLDERS
# -----------------------------
SETTINGS_PATH = os.path.join("TEAM_CONFIG", "team_settings.json")
ASSETS_DIR = "assets"
os.makedirs(ASSETS_DIR, exist_ok=True)

# FORCE include team data folders (Streamlit Cloud quirk) — but don't crash if missing
try:
    if os.path.exists("data/teams"):
        _ = os.listdir("data/teams")
except Exception:
    pass


# -----------------------------
# SETTINGS LOADER
# -----------------------------
def load_settings():
    defaults = {
        "app_title": "RP Spray Charts",
        "subtitle": "Coaches that want to win, WILL put in the time",
        "primary_color": "#b91c1c",
        "secondary_color": "#111111",
        "background_image": os.path.join("assets", "background.jpg"),
        "logo_image": os.path.join("assets", "logo.png"),
        "strict_mode_default": True,
    }
    if os.path.exists(SETTINGS_PATH):
        try:
            with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
                user = json.load(f)
            if isinstance(user, dict):
                defaults.update({k: v for k, v in user.items() if v is not None})
        except Exception:
            pass
    return defaults


SETTINGS = load_settings()
settings = SETTINGS  # alias so the rest of the code can use `settings`


# -----------------------------
# ✅ MUST BE FIRST STREAMLIT CALL
# -----------------------------
st.set_page_config(
    page_title=SETTINGS.get("app_title", "RP Spray Charts"),
    page_icon="⚾",
    layout="wide",
)

# ============================
# ACCESS CODE GATE (CLEAN + STABLE)
# ============================

from datetime import datetime, timezone

@st.cache_data(show_spinner=False)
def load_team_codes() -> dict:
    """
    Loads active teams from Supabase.
    Returns a dict keyed by team_code (UPPER), value = row dict.
    NOTE: We intentionally do NOT key by slug to avoid collisions / confusion.
    """
    try:
        res = (
            supabase.table("team_access")
            .select("id, team_slug, team_code, team_name, code_hash, is_active")
            .eq("is_active", True)
            .execute()
        )
        rows = res.data or []
        out = {}
        for r in rows:
            code = str(r.get("team_code") or "").strip().upper()
            if code:
                out[code] = r
        return out
    except Exception:
        return {}


def license_is_active(team_code: str) -> bool:
    """
    Returns True if team has active license (and not expired if expires_at set).
    Table: licenses (team_code text, status text, expires_at timestamptz)
    """
    try:
        tc = str(team_code or "").strip().upper()
        if not tc:
            return False

        res = (
            supabase.table("licenses")
            .select("status, expires_at")
            .eq("team_code", tc)
            .limit(1)
            .execute()
        )
        rows = res.data or []
        if not rows:
            return False

        row = rows[0] or {}
        status = str(row.get("status", "")).strip().lower()
        if status != "active":
            return False

        exp = row.get("expires_at")
        if exp:
            exp_dt = datetime.fromisoformat(str(exp).replace("Z", "+00:00"))
            if exp_dt < datetime.now(timezone.utc):
                return False

        return True
    except Exception:
        return False


def require_team_access():
    # ---------------------------------
    # Already unlocked?
    # ---------------------------------
    team_code = str(st.session_state.get("team_code", "") or "").strip().upper()
    if team_code:
        return team_code, {"team_code": team_code}

    st.markdown("## Enter Access Code")

    code_raw = st.text_input(
        "Access Code",
        type="password",
        placeholder="Enter Access Code",
        key="access_code_input",
)


    # ---------------------------------
    # NORMAL UNLOCK
    # ---------------------------------
    if st.button("Unlock", key="unlock_btn"):
        entered = (code_raw or "").strip()

        if not entered:
            st.error("Enter an access code")
            st.stop()

        entered_hash = hash_access_code(entered)

        res = (
            supabase.table("team_access")
            .select("team_code, code_hash")
            .eq("is_active", True)
            .execute()
        )

        rows = res.data or []
        matched = None
        for r in rows:
            stored = str((r or {}).get("code_hash", "")).strip()
            if stored and entered_hash == stored:
                matched = r
                break

        if not matched:
            st.error("Invalid access code")
            st.stop()

        team_code = str(matched.get("team_code", "") or "").strip().upper()

        if not license_is_active(team_code):
            st.error("License inactive. Contact admin.")
            st.stop()

        st.session_state.team_code = team_code
        st.rerun()

    st.stop()




TEAM_CODE, _ = require_team_access()


# -----------------------------
# TEAM CFG LOADER (FILE)
# -----------------------------
def _load_team_cfg_from_file(team_code: str) -> dict:
    try:
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)

        teams = data.get("teams", {}) or {}
        branding = data.get("team_branding", {}) or {}

        cfg = None
        for _, t in teams.items():
            if str(t.get("team_code", "")).strip().upper() == str(team_code).strip().upper():
                cfg = t
                break

        cfg = cfg or {}

        b = branding.get(str(team_code).strip().upper(), {}) or {}
        if b.get("logo_path"):
            cfg["logo_path"] = b["logo_path"]
        if b.get("background_path"):
            cfg["background_path"] = b["background_path"]

        return cfg
    except Exception:
        return {}


TEAM_CFG = _load_team_cfg_from_file(TEAM_CODE) or {}

# ===============================
# TERMS OF USE — HARD GATE (PAGE-LEVEL)
# ===============================
_TERMS_KEY = f"terms_accepted__{str(TEAM_CODE).strip().upper()}"
_AGREE_KEY = f"terms_agree__{str(TEAM_CODE).strip().upper()}"

if _TERMS_KEY not in st.session_state:
    st.session_state[_TERMS_KEY] = False

if not st.session_state[_TERMS_KEY]:

    st.title("Terms of Use")

    terms_text = """
TERMS OF USE — RP SPRAY ANALYTICS

By accessing or using RP Spray Analytics, you acknowledge and agree to the following:

1. LICENSED USE (NOT SOLD)
RP Spray Analytics is licensed, not sold. Each license grants one baseball program (school) a non-exclusive, non-transferable right to use the application for internal team and coaching purposes during the licensed season only.

2. PRICING & SCOPE
Licenses are issued per program, per season. One license covers unlimited games, players, and exports for the licensed program during the active season.

3. AUTHORIZED USERS
Access is restricted to authorized coaches and staff of the licensed program. Access codes may not be shared, transferred, or reused by another team, individual, or organization.

4. PROHIBITED USE
• Unauthorized copying, redistribution, resale, sublicensing, or public sharing
• Reverse engineering, decompilation, or replication of logic or outputs
• Competitive, commercial, or third-party use

5. DATA OWNERSHIP
Teams retain ownership of raw data. All analytics, workflows, and outputs remain proprietary.

6. TERMINATION
Access may be revoked immediately for violations without refund.
"""

    st.markdown(
        f"""
        <div style="
            height: 360px;
            overflow-y: auto;
            padding: 16px;
            border: 1px solid #d1d5db;
            border-radius: 8px;
            background-color: #ffffff;
            font-size: 0.95rem;
        ">
            <pre style="white-space: pre-wrap; margin: 0;">{terms_text}</pre>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.form(key=f"terms_form__{str(TEAM_CODE).strip().upper()}"):
        st.checkbox("I have read and agree to the Terms of Use", key=_AGREE_KEY)
        submitted = st.form_submit_button("Continue")

    if submitted:
        if st.session_state.get(_AGREE_KEY, False):
            st.session_state[_TERMS_KEY] = True
            st.rerun()
        else:
            st.warning("You must agree before continuing.")

    st.stop()  # ✅ ONLY stop while locked

# -----------------------------
# RESOLVED TEAM BRANDING (logo + background)
# -----------------------------
LOGO_PATH = TEAM_CFG.get("logo_path") or SETTINGS.get("logo_image")
BG_PATH   = TEAM_CFG.get("background_path") or SETTINGS.get("background_image")


# -----------------------------
# ✅ TEAM-ISOLATED STORAGE (folders only for rosters; totals are in Supabase)
# -----------------------------
TEAM_CODE_SAFE = str(TEAM_CODE).strip().upper()
TEAM_ROOT = os.path.join("data", "teams", TEAM_CODE_SAFE)
TEAM_ROSTERS_DIR = os.path.join(TEAM_ROOT, "rosters")
TEAM_SEASON_DIR = os.path.join(TEAM_ROOT, "season_totals")  # legacy folder; not used for season totals anymore
os.makedirs(TEAM_ROSTERS_DIR, exist_ok=True)
os.makedirs(TEAM_SEASON_DIR, exist_ok=True)


# -----------------------------
# ENGINE CONSTANTS (MUST EXIST BEFORE empty_stat_dict/db_load)
# -----------------------------
# ✅ Bunts combined into ONE stat (Sac + regular) and kept separate from GB/FB
LOCATION_KEYS = ["LF", "CF", "RF", "3B", "SS", "2B", "1B", "P"]

BALLTYPE_KEYS = ["GB", "FB"]

# Combo keys only for true field locations (NO BUNT/UNKNOWN)
COMBO_LOCS = [loc for loc in LOCATION_KEYS if loc not in ["BUNT", "UNKNOWN"]]
COMBO_KEYS = [f"GB-{loc}" for loc in COMBO_LOCS] + [f"FB-{loc}" for loc in COMBO_LOCS]

# ✅ BASERUNNING RE-ENABLED (NO SB-H / CS-H)
RUN_KEYS = ["SB", "SB-2B", "SB-3B", "CS", "CS-2B", "CS-3B"]

# Games Played tracking (per player)
GP_KEY = "GP"
BUNTS_KEY = "Bunts"



# -----------------------------
# STAT HELPERS
# -----------------------------
def empty_stat_dict():
    d = {loc: 0 for loc in LOCATION_KEYS}
    for k in BALLTYPE_KEYS:
        d[k] = 0
    for ck in COMBO_KEYS:
        d[ck] = 0
    for rk in RUN_KEYS:
        d[rk] = 0
    d[GP_KEY] = 0
    d[BUNTS_KEY] = 0
    return d


def ensure_all_keys(d: dict):
    for loc in LOCATION_KEYS:
        d.setdefault(loc, 0)
    for k in BALLTYPE_KEYS:
        d.setdefault(k, 0)
    for ck in COMBO_KEYS:
        d.setdefault(ck, 0)
    for rk in globals().get("RUN_KEYS", []):
        d.setdefault(rk, 0)
    d.setdefault(GP_KEY, 0)
    d.setdefault(BUNTS_KEY, 0)
    return d


# -----------------------------
# PBP NORMALIZATION + GAME HASH
# -----------------------------
def normalize_pbp(text: str) -> str:
    return "\n".join([ln.strip() for ln in (text or "").strip().splitlines() if ln.strip()])


def game_key_from_pbp(team_key: str, pbp_text: str) -> str:
    norm = normalize_pbp(pbp_text)
    h = hashlib.sha1((team_key + "||" + norm).encode("utf-8")).hexdigest()
    return f"pbp_sha1_{h}"


# -----------------------------
# REGEX / PATTERNS (ENGINE)
# -----------------------------
GB_REGEX = [
    re.compile(r"\bground(?:s|ed)?\b"),
    re.compile(r"\bground ?ball\b"),
    re.compile(r"\bgrounder\b"),
    re.compile(r"\bchopper\b"),
    re.compile(r"\bbouncer\b"),
    re.compile(r"\bdribbler\b"),
    re.compile(r"\broller\b"),
    re.compile(r"\btapper\b"),
    re.compile(r"\bslow[- ]roller\b"),
]
LD_REGEX = [
    re.compile(r"\bline drive\b"),
    re.compile(r"\blines?\b"),
    re.compile(r"\blined\b"),
    re.compile(r"\bon a line\b"),
]
FB_REGEX = [
    re.compile(r"\bfly(?:\s?ball)?\b"),
    re.compile(r"\bflies\b"),
    re.compile(r"\bflied\b"),
    re.compile(r"\bpops?\b"),
    re.compile(r"\bpop[- ]?up\b"),
    re.compile(r"\bpopup\b"),
    re.compile(r"\btowering fly\b"),
    re.compile(r"\bhigh fly\b"),
    re.compile(r"\bdeep fly\b"),
    re.compile(r"\bshallow fly\b"),
    re.compile(r"\binfield fly\b"),
    re.compile(r"\bfoul pop\b"),
    re.compile(r"\bblooper\b"),
    re.compile(r"\bflare\b"),
    re.compile(r"\bfloater\b"),
    re.compile(r"\blofted\b"),
]
SACFLY_REGEX = [re.compile(r"\bsac(?:rifice)? fly\b")]

LF_PATTERNS = [
    "left fielder ", "to left fielder", "to left field", "to left", "into left field",
    "down the left field line", "down the left-field line",
    "down the lf line", "down the left line", "toward left field",
    "into shallow left", "into deep left", "into left-center", "into left center",
    "in front of left fielder"
]
CF_PATTERNS = [
    "center fielder ", "to center fielder", "to center field", "to center", "into center field",
    "into deep center", "into shallow center",
    "into left-center field", "into left center field",
    "into right-center field", "into right center field",
    "up the middle into center", "up the middle to center"
]
RF_PATTERNS = [
    "right fielder ", "to right fielder", "to right field", "to right", "into right field",
    "down the right field line", "down the right-field line",
    "down the rf line", "toward right field",
    "into shallow right", "into deep right",
    "into right-center", "into right center",
    "in front of right fielder"
]
SS_PATTERNS = [
    "shortstop ", "to shortstop", "to the shortstop", "to ss",
    "fielded by the shortstop", "fielded by shortstop",
    "shortstop fields", "shortstop to", "shortstop throws to",
    "shortstop makes the play", "at shortstop"
]
_2B_PATTERNS = [
    "second baseman ", "to second baseman", "to the second baseman", "to 2nd baseman",
    "fielded by second baseman", "fielded by the second baseman",
    "second baseman fields", "second baseman to", "second baseman throws to"
]
_3B_PATTERNS = [
    "third baseman ", "to third baseman", "to the third baseman", "to 3rd baseman",
    "fielded by third baseman", "fielded by the third baseman",
    "third baseman fields", "third baseman to", "third baseman throws to"
]
_1B_PATTERNS = [
    "first baseman ", "to first baseman", "to the first baseman", "to 1st baseman",
    "fielded by first baseman", "fielded by the first baseman",
    "first baseman fields", "first baseman to", "first baseman throws to"
]
P_PATTERNS = [
    "to pitcher", "to the pitcher", "back to the pitcher",
    "back to pitcher", "back to the mound",
    "fielded by pitcher", "fielded by the pitcher",
    "pitcher fields", "pitcher to", "pitcher throws to",
    "back up the middle to the pitcher"
]
LEFT_SIDE_PATTERNS = [
    "through the left side", "up the left side", "toward the left side",
    "between shortstop and third baseman", "between 3rd and ss",
    "between ss and 3b", "between short and third"
]
RIGHT_SIDE_PATTERNS = [
    "through the right side", "up the right side", "toward the right side",
    "between first baseman and second baseman", "between 1st and 2nd",
    "between second and first"
]


PAREN_NAME_REGEX = re.compile(r"\(([^)]+)\)")

# -----------------------------
# SB / CS REGEX (STRICT + CLEAN)
# -----------------------------
SB_ACTION_REGEX = re.compile(r"\b(steals?|stole)\s+(2nd|second|3rd|third)\b", re.IGNORECASE)
CS_ACTION_REGEX = re.compile(r"\b(caught\s+stealing|out\s+stealing)\s+(2nd|second|3rd|third)\b", re.IGNORECASE)

def extract_runner_before_index(line: str, idx: int, roster: set[str]) -> Optional[str]:
    """
    Finds the runner name to the LEFT of the steals/CS phrase.
    Uses roster longest-match-first for 98%+ accuracy.
    """
    if not line or idx is None:
        return None

    left = line[:idx]
    chunk = left.split(",")[-1].strip()
    chunk = re.sub(r"\([^)]*\)", "", chunk)
    chunk = re.sub(r"\s+", " ", chunk).strip()
    if not chunk:
        return None

    roster_sorted = sorted((r.strip().strip('"') for r in roster if r), key=len, reverse=True)
    chunk_lower = chunk.lower()

    for rn in roster_sorted:
        if rn and chunk_lower.endswith(rn.lower()):
            return rn

    parts = chunk.split()
    if len(parts) >= 2:
        cand = parts[-2] + " " + parts[-1]
        if cand in roster:
            return cand

    return None


def normalize_base_bucket(prefix: str, base_raw: Optional[str]) -> str:
    if not base_raw:
        return prefix
    b = base_raw.strip().lower().strip("()").strip()
    if b in ["2nd", "second"]:
        return f"{prefix}-2B"
    if b in ["3rd", "third"]:
        return f"{prefix}-3B"
    if b == "home":
        return prefix  # ✅ we do NOT track -H buckets
        
    return prefix


BAD_FIRST_TOKENS = {
    "top", "bottom", "inning", "pitch", "ball", "strike", "foul",
    "runner", "runners", "advances", "advance", "steals", "stole", "caught",
    "substitution", "defensive", "offensive", "double", "triple", "single", "home",
    "out", "safe", "error", "no", "one", "two", "three",
}


def starts_like_name(token: str) -> bool:
    if not token:
        return False
    t = token.strip().strip('"').strip().lower()
    return t[:1].isalpha() and t not in BAD_FIRST_TOKENS


def overall_confidence_score(conf_val: int):
    if conf_val >= 4:
        return "high"
    if conf_val >= 2:
        return "medium"
    return "low"


def get_batter_name(line: str, roster: set[str]):
    line = (line or "").strip().strip('"')
    if not line:
        return None

    # Remove parenthetical junk and normalize spacing
    clean = re.sub(r"\([^)]*\)", "", line)
    clean = re.sub(r"\s+", " ", clean).strip()
    if not clean:
        return None

    parts = clean.split()
    if not parts:
        return None

    # First token must look like a name starter
    if not starts_like_name(parts[0]):
        return None

    # 🔥 PRIMARY MATCH: longest roster name that matches the start of the line
    roster_sorted = sorted(
        (r.strip().strip('"') for r in roster if r),
        key=len,
        reverse=True
    )

    for rname in roster_sorted:
        if clean == rname or clean.startswith(rname + " "):
            return rname

    # Fallback: first initial + last token
    if len(parts) >= 2:
        candidate = parts[0] + " " + parts[-1]
        if candidate in roster:
            return candidate

    return None


def extract_runner_name_fallback(clean_line: str, roster: set[str]) -> Optional[str]:
    runner = get_batter_name(clean_line, roster)
    if runner:
        return runner

    pm = PAREN_NAME_REGEX.search(clean_line)
    if pm:
        inside = re.sub(r"\s+", " ", pm.group(1).strip())
        runner = get_batter_name(inside, roster)
        if runner:
            return runner

    return None


def parse_running_event(clean_line: str, roster: set[str]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Returns (runner_name, total_key, base_key) or (None, None, None).
    ✅ SB/CS are counted ONLY when a runner is confidently identified from roster.
    ✅ No SB-H / CS-H ever produced.
    """
    line = (clean_line or "").strip()
    if not line:
        return None, None, None

    # SB
    m = SB_ACTION_REGEX.search(line)
    if m:
        base_key = normalize_base_bucket("SB", m.group(2))
        runner = extract_runner_before_index(line, m.start(), roster) or extract_runner_name_fallback(line, roster)
        if runner:
            return runner, "SB", base_key
        return None, None, None

    # CS
    m = CS_ACTION_REGEX.search(line)
    if m:
        base_key = normalize_base_bucket("CS", m.group(2))
        runner = extract_runner_before_index(line, m.start(), roster) or extract_runner_name_fallback(line, roster)
        if runner:
            return runner, "CS", base_key
        return None, None, None

    return None, None, None


def is_ball_in_play(line_lower: str) -> bool:
    ll = (line_lower or "").strip()
    if not ll:
        return False

    # exclude non-BIP and running events
    if any(kw in ll for kw in [
        "hit by pitch", "hit-by-pitch", "hit batsman",
        "walks", "walked", " base on balls", "intentional walk",
        "strikes out", "strikeout", "called out on strikes",
        "reaches on catcher interference", "catcher's interference",
        "defensive indifference",
        "picked off", "pickoff",
    ]):
        return False

    bip_outcomes = [
        "grounds", "grounded", "ground ball", "groundball", "grounder",
        "singles", "doubles", "triples", "homers", "home run",
        "lines out", "line drive", "lined out", "line out",
        "flies out", "fly ball", "flied out", "fly out",
        "pops out", "pop up", "pop-out", "popup",
        "bloops", "blooper",
        "bunts", "bunt", "sacrifice bunt", "sac bunt", "sacrifice hit",
        "sac fly", "sacrifice fly",
        "reaches on a fielding error", "reaches on a throwing error",
        "reaches on error", "reached on error", "safe on error",
        "reaches on a missed catch error",
        "fielder's choice", "fielders choice",
        "double play", "triple play",
        "out at first", "out at second", "out at third", "out at home",
    ]
    if any(kw in ll for kw in bip_outcomes):
        return True

    # fallback: any explicit fielder/location markers
    fielder_markers = [
        "left fielder", "center fielder", "right fielder",
        "shortstop", "second baseman", "third baseman", "first baseman",
        "to left field", "to center field", "to right field",
        "to shortstop", "to second baseman", "to third baseman", "to first baseman",
        "to pitcher", "back to the mound",
        "down the left", "down the right", "left-center", "right-center"
    ]
    return any(m in ll for m in fielder_markers)


def classify_ball_type(line_lower: str):
    # ✅ Bunts are NOT GB/FB in this system
    if "bunt" in line_lower:
        return None, 3, ["Bunt detected → no GB/FB"]

    for rx in SACFLY_REGEX:
        if rx.search(line_lower):
            return "FB", 3, ["Matched sac fly regex → FB"]

    for rx in LD_REGEX:
        if rx.search(line_lower):
            return "FB", 2, ["Matched line drive regex → FB"]

    for rx in GB_REGEX:
        if rx.search(line_lower):
            return "GB", 2, [f"Matched GB regex: {rx.pattern}"]

    for rx in FB_REGEX:
        if rx.search(line_lower):
            return "FB", 2, [f"Matched FB regex: {rx.pattern}"]

    return None, 0, []


def classify_location(line_lower: str, strict_mode: bool = False):

    # ✅ Any bunt type: do NOT return a location (we count it separately as "Bunts")
    if "sacrifice bunt" in line_lower or "sac bunt" in line_lower or "sacrifice hit" in line_lower:
        return None, 3, ["Sac bunt detected → Bunts stat only"]

    # ✅ Any other bunt: also no location
    if "bunt" in line_lower:
        return None, 3, ["Bunt detected → Bunts stat only"]

    candidates = []
    
    def add_candidates(patterns, code, label):
        for kw in patterns:
            idx = line_lower.find(kw)
            if idx != -1:
                candidates.append((idx, code, f"Matched {label} phrase: '{kw}'"))

    add_candidates(LF_PATTERNS, "LF", "LF")
    add_candidates(CF_PATTERNS, "CF", "CF")
    add_candidates(RF_PATTERNS, "RF", "RF")
    add_candidates(SS_PATTERNS, "SS", "SS")
    add_candidates(_3B_PATTERNS, "3B", "3B")
    add_candidates(_2B_PATTERNS, "2B", "2B")
    add_candidates(_1B_PATTERNS, "1B", "1B")
    add_candidates(P_PATTERNS, "P", "P")

    if candidates:
        _, loc, reason = min(candidates, key=lambda x: x[0])
        return loc, 3, [reason]

    if strict_mode:
        return None, 0, ["Strict mode: no explicit fielder/location phrase found"]

    for kw in LEFT_SIDE_PATTERNS:
        if kw in line_lower:
            return "SS", 1, [f"Matched left-side phrase: '{kw}' → approximate SS"]

    for kw in RIGHT_SIDE_PATTERNS:
        if kw in line_lower:
            return "2B", 1, [f"Matched right-side phrase: '{kw}' → approximate 2B"]

    return None, 0, []


# -----------------------------
# UNLIMITED TEAMS: read roster files
# -----------------------------
def list_team_files():
    files = []
    for fn in os.listdir(TEAM_ROSTERS_DIR):
        if fn.lower().endswith(".txt"):
            files.append(fn)
    files.sort(key=lambda x: x.lower())
    return files


def team_name_from_file(filename: str) -> str:
    return os.path.splitext(filename)[0]


def safe_team_key(team_name: str) -> str:
    key = re.sub(r"[^a-zA-Z0-9]+", "_", team_name.strip()).strip("_").lower()
    return key or "team"


def roster_path_for_file(filename: str) -> str:
    return os.path.join(TEAM_ROSTERS_DIR, filename)


def load_roster_text(path: str) -> str:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    return ""


def save_roster_text(path: str, text: str):
    with open(path, "w", encoding="utf-8") as f:
        f.write(text.strip() + "\n" if text.strip() else "")


def add_game_to_season(season_team, season_players, game_team, game_players):
    # ✅ include Bunts in roll-up
    KEYS = LOCATION_KEYS + BALLTYPE_KEYS + COMBO_KEYS + RUN_KEYS + [GP_KEY, BUNTS_KEY]

    for key in KEYS:
        season_team[key] = season_team.get(key, 0) + game_team.get(key, 0)

    for player, gstats in game_players.items():
        season_players.setdefault(player, empty_stat_dict())
        sstats = season_players[player]
        for key in KEYS:
            sstats[key] = sstats.get(key, 0) + gstats.get(key, 0)


# -----------------------------
# SUPABASE (persistent storage)
# -----------------------------
SUPABASE_SETUP_SQL = """
-- ============================
-- SEASON TOTALS (one row per team_code + team_key)
-- ============================
create table if not exists public.season_totals (
  id bigserial primary key,
  team_code text not null,
  team_key  text not null,
  data jsonb not null default '{}'::jsonb,
  games_played integer not null default 0,
  updated_at timestamptz not null default now()
);

create unique index if not exists season_totals_unique
  on public.season_totals (team_code, team_key);

create index if not exists season_totals_team_idx
  on public.season_totals (team_code, team_key);
  -- ============================
-- TEAM ROSTERS (persistent, per team_code)
-- ============================
create table if not exists public.team_rosters (
  id bigserial primary key,
  team_code   text not null,
  team_key    text not null,
  team_name   text not null,
  roster_text text not null default '',
  updated_at  timestamptz not null default now()
);

create unique index if not exists team_rosters_unique
  on public.team_rosters (team_code, team_key);

create index if not exists team_rosters_team_idx
  on public.team_rosters (team_code, team_name);


-- ============================
-- PROCESSED GAMES (hard dedupe)
-- ============================
create table if not exists public.processed_games (
  id bigserial primary key,
  team_code text not null,
  team_key  text not null,
  game_hash text not null,
  created_at timestamptz not null default now()
);

create unique index if not exists processed_games_unique
  on public.processed_games (team_code, team_key, game_hash);

create index if not exists processed_games_team_idx
  on public.processed_games (team_code, team_key);
""".strip()


def _show_db_error(e: Exception, label: str):
    st.error(f"**{label}**")
    try:
        parts = [f"type: {type(e)}", f"error: {str(e)}"]
        for attr in ("message", "details", "hint", "code"):
            if hasattr(e, attr):
                val = getattr(e, attr)
                if val:
                    parts.append(f"{attr}: {val}")
        st.code("\n".join(parts), language="text")

        # ✅ show full traceback so we get the exact line that caused it
        st.code(traceback.format_exc(), language="text")
    except Exception:
        st.write(str(e))



def _render_supabase_fix_block():
    st.error("Supabase tables are missing or mismatched (season_totals / processed_games).")
    ("### Fix (copy/paste into Supabase → SQL Editor → Run)")
    st.code(SUPABASE_SETUP_SQL, language="sql")
    (
        """
**Then refresh your Streamlit app.**  
If it still errors after running the SQL, your Streamlit **secrets** are wrong.
"""
    )


@st.cache_resource(show_spinner=False)
def get_supabase() -> Client:
    url = st.secrets.get("SUPABASE_URL", "").strip()
    key = st.secrets.get("SUPABASE_SERVICE_KEY", "").strip()  # service role key (server-side only)
    if not url or not key:
        raise RuntimeError("Missing SUPABASE_URL or SUPABASE_SERVICE_KEY in Streamlit secrets.")
    return create_client(url, key)


try:
    supabase = get_supabase()
except Exception as e:
    _show_db_error(e, "Supabase secrets missing / invalid")
    st.stop()


def supa_execute_with_retry(builder, tries: int = 5):
    last_err = None
    for i in range(tries):
        try:
            return builder.execute()
        except (httpx.ReadError, httpx.ConnectError, httpx.ReadTimeout) as e:
            last_err = e
            time.sleep(0.6 * (i + 1))
    raise last_err


def supabase_health_check_or_stop():
    try:
        supa_execute_with_retry(supabase.table("season_totals").select("id").limit(1))
        supa_execute_with_retry(supabase.table("processed_games").select("id").limit(1))
        supa_execute_with_retry(supabase.table("team_rosters").select("team_code").limit(1))


        return True
    except Exception as e:
        _show_db_error(e, "Supabase not ready")
        _render_supabase_fix_block()
        st.stop()


supabase_health_check_or_stop()


def db_load_season_totals(team_code: str, team_key: str, current_roster: set[str]):
    """
    Returns (season_team, season_players, games_played, processed_hashes_set, archived_players_set)
    archived_players are players who exist in DB totals but are not on the current roster (or were removed).
    """
    season_team = empty_stat_dict()
    season_players = {p: empty_stat_dict() for p in current_roster}
    games_played = 0
    archived_players = set()

    try:
        res = (
            supabase.table("season_totals")
            .select("data, games_played")
            .eq("team_code", team_code)
            .eq("team_key", team_key)
            .limit(1)
            .execute()
        )
    except Exception as e:
        _show_db_error(e, "Supabase SELECT failed on season_totals")
        _render_supabase_fix_block()
        st.stop()

    if res.data:
        row = res.data[0]
        payload = row.get("data") or {}

        raw_team = payload.get("team") or {}
        raw_players = payload.get("players") or {}
        raw_meta = payload.get("meta") or {}

        season_team = ensure_all_keys(raw_team if isinstance(raw_team, dict) else {})
        season_players = {}

        if isinstance(raw_players, dict):
            for p, stats in raw_players.items():
                season_players[p] = ensure_all_keys(stats) if isinstance(stats, dict) else empty_stat_dict()

        # Ensure current roster always exists in dict (so new guys show up immediately)
        for p in current_roster:
            if p not in season_players:
                season_players[p] = empty_stat_dict()

        games_played = int(row.get("games_played") or 0)

        # Archived list stored in meta (optional)
        ap = raw_meta.get("archived_players", []) if isinstance(raw_meta, dict) else []
        if isinstance(ap, list):
            archived_players = {str(x).strip().strip('"') for x in ap if str(x).strip()}

    try:
        pres = (
            supabase.table("processed_games")
            .select("game_hash")
            .eq("team_code", team_code)
            .eq("team_key", team_key)
            .execute()
        )
    except Exception as e:
        _show_db_error(e, "Supabase SELECT failed on processed_games")
        _render_supabase_fix_block()
        st.stop()

    processed_set = set()
    if pres.data:
        processed_set = {r["game_hash"] for r in pres.data if r.get("game_hash")}

    return season_team, season_players, games_played, processed_set, archived_players




def db_get_coach_notes(team_code: str, team_key: str) -> str:
    """Fetch per-opponent coach notes from season_totals.data.meta.coach_notes."""
    try:
        res = (
            supabase.table("season_totals")
            .select("data")
            .eq("team_code", team_code)
            .eq("team_key", team_key)
            .limit(1)
            .execute()
        )
        if res.data:
            payload = res.data[0].get("data") or {}
            meta = payload.get("meta") or {}
            if isinstance(meta, dict):
                return str(meta.get("coach_notes", "") or "")
        return ""
    except Exception:
        return ""

def db_get_player_notes(team_code: str, team_key: str) -> str:
    """Fetch per-player coach notes from season_totals.data.meta.player_notes (JSON string)."""
    try:
        res = (
            supabase.table("season_totals")
            .select("data")
            .eq("team_code", team_code)
            .eq("team_key", team_key)
            .limit(1)
            .execute()
        )
        if res.data:
            payload = res.data[0].get("data") or {}
            meta = payload.get("meta") or {}
            if isinstance(meta, dict):
                return str(meta.get("player_notes", "") or "")
        return ""
    except Exception:
        return ""

def db_save_season_totals(
    team_code: str,
    team_key: str,
    season_team: dict,
    season_players: dict,
    games_played: int,
    archived_players: set[str] | list[str] | None = None,
    coach_notes: str | None = None,
    player_notes: str | None = None,
):
    archived_list = []
    if archived_players:
        archived_list = sorted({str(x).strip().strip('"') for x in archived_players if str(x).strip()})

    # Preserve existing meta so roster/game saves don't wipe notes
    existing_meta: dict = {}
    try:
        res0 = (
            supabase.table("season_totals")
            .select("data")
            .eq("team_code", team_code)
            .eq("team_key", team_key)
            .limit(1)
            .execute()
        )
        if res0.data:
            payload0 = res0.data[0].get("data") or {}
            meta0 = payload0.get("meta") or {}
            if isinstance(meta0, dict):
                existing_meta = dict(meta0)
    except Exception:
        existing_meta = {}

    existing_meta["archived_players"] = archived_list
    if coach_notes is not None:
        existing_meta["coach_notes"] = str(coach_notes)

    payload = {
        "team": season_team,
        "players": season_players,
        "meta": existing_meta,
    }

    try:
        (
            supabase.table("season_totals")
            .upsert(
                {
                    "team_code": team_code,
                    "team_key": team_key,
                    "data": payload,
                    "games_played": int(games_played),
                    "updated_at": datetime.utcnow().isoformat(),
                },
                on_conflict="team_code,team_key",
            )
            .execute()
        )
    except Exception as e:
        _show_db_error(e, "Supabase UPSERT failed on season_totals")
        _render_supabase_fix_block()
        st.stop()


def db_try_mark_game_processed(team_code: str, team_key: str, game_hash: str) -> bool:
    try:
        supabase.table("processed_games").insert(
            {"team_code": team_code, "team_key": team_key, "game_hash": game_hash}
        ).execute()
        return True
    except Exception:
        return False


def db_unmark_game_processed(team_code: str, team_key: str, game_hash: str):
    try:
        supa_execute_with_retry(
            supabase.table("processed_games")
            .delete()
            .eq("team_code", team_code)
            .eq("team_key", team_key)
            .eq("game_hash", game_hash)
        )
    except Exception:
        pass
        
def db_reset_season(team_code: str, team_key: str):
    try:
        supabase.table("season_totals").delete().eq("team_code", team_code).eq("team_key", team_key).execute()
        supabase.table("processed_games").delete().eq("team_code", team_code).eq("team_key", team_key).execute()
    except Exception as e:
        _show_db_error(e, "Supabase RESET failed")
        _render_supabase_fix_block()
        st.stop()

      import time
import httpx
from datetime import datetime

# -----------------------------
# SUPABASE EXECUTE (RETRY)
# -----------------------------
def _sb_execute(q, tries: int = 3, base_sleep: float = 0.4):
    """
    Retry wrapper for Supabase/PostgREST calls to reduce transient httpx/httpcore ReadError.
    """
    last_err = None
    for i in range(tries):
        try:
            return q.execute()
        except (httpx.ReadError, httpx.ConnectError, httpx.RemoteProtocolError) as e:
            last_err = e
            time.sleep(base_sleep * (i + 1))
        except Exception as e:
            # still retry a couple times — Streamlit reruns can collide
            last_err = e
            time.sleep(base_sleep * (i + 1))
    raise last_err


# -----------------------------
# DB: TEAM ROSTERS
# -----------------------------
@st.cache_data(ttl=30, show_spinner=False)
def db_list_teams(team_code: str):
    """
    Returns list of dicts: [{team_key, team_name, roster_text, updated_at}]
    Cached briefly to prevent hammering Supabase on reruns.
    """
    try:
        res = _sb_execute(
            supabase.table("team_rosters")
            .select("team_key, team_name, roster_text, updated_at")
            .eq("team_code", str(team_code).strip().upper())
            .order("team_name")
        )
        return res.data or []
    except Exception as e:
        _show_db_error(e, "Supabase SELECT failed on team_rosters")
        _render_supabase_fix_block()
        st.stop()


def db_get_team(team_code: str, team_key: str):
    try:
        res = _sb_execute(
            supabase.table("team_rosters")
            .select("team_key, team_name, roster_text, updated_at")
            .eq("team_code", str(team_code).strip().upper())
            .eq("team_key", str(team_key).strip())
            .limit(1)
        )
        if res.data:
            return res.data[0]
        return None
    except Exception as e:
        _show_db_error(e, "Supabase SELECT failed on team_rosters (single)")
        _render_supabase_fix_block()
        st.stop()


def db_get_roster(team_code: str, team_key: str) -> str:
    """
    Returns roster_text for a team, or empty string if none exists.
    """
    team = db_get_team(team_code, team_key)
    if team and team.get("roster_text"):
        return team["roster_text"]
    return ""


def db_upsert_team(team_code: str, team_key: str, team_name: str, roster_text: str):
    """
    Upserts one roster row per (team_code, team_key).
    Requires unique(team_code, team_key) on team_rosters.
    Clears cache so UI updates immediately.
    """
    try:
        _sb_execute(
            supabase.table("team_rosters")
            .upsert(
                {
                    "team_code": str(team_code).strip().upper(),
                    "team_key": str(team_key).strip(),
                    "team_name": str(team_name or "").strip(),
                    "roster_text": roster_text or "",
                    "updated_at": datetime.utcnow().isoformat(),
                },
                on_conflict="team_code,team_key",
            )
        )
        db_list_teams.clear()  # bust cache after write
    except Exception as e:
        _show_db_error(e, "Supabase UPSERT failed on team_rosters")
        _render_supabase_fix_block()
        st.stop()


def db_delete_team(team_code: str, team_key: str):
    """
    Optional: delete a team roster row.
    Clears cache so UI updates immediately.
    """
    try:
        _sb_execute(
            supabase.table("team_rosters")
            .delete()
            .eq("team_code", str(team_code).strip().upper())
            .eq("team_key", str(team_key).strip())
        )
        db_list_teams.clear()  # bust cache after delete
    except Exception as e:
        _show_db_error(e, "Supabase DELETE failed on team_rosters")
        _render_supabase_fix_block()
        st.stop()

  

# -----------------------------
# BRANDING + BACKGROUND
# -----------------------------
def get_base64_image(path: str) -> str:
    if not path or not os.path.exists(path):
        return ""
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


PRIMARY = SETTINGS.get("primary_color", "#b91c1c")
SECONDARY = SETTINGS.get("secondary_color", "#111111")

LOGO_PATH = (
    TEAM_CFG.get("logo_path")
    or SETTINGS.get("logo_image")
    or os.path.join("assets", "logo.png")
)

BG_PATH = (
    TEAM_CFG.get("background_path")
    or SETTINGS.get("background_image")
    or os.path.join("assets", "background.jpg")
)

if TEAM_CFG:
    LOGO_PATH = TEAM_CFG.get("logo_path", LOGO_PATH)
    BG_PATH = TEAM_CFG.get("background_path", BG_PATH)

BG_B64 = get_base64_image(BG_PATH)


# -----------------------------
# FONTS (force load)
# -----------------------------
st.markdown(
    """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Black+Ops+One&family=Jersey+10&display=swap" rel="stylesheet">
""",
    unsafe_allow_html=True,
)

# -----------------------------
# STYLES (sharp + bordered title)
# -----------------------------
st.markdown(
    f"""
<style>
h1.app-title {{
    font-family: 'Black Ops One', 'Jersey 10', sans-serif !important;
    font-size: 4.0rem !important;
    color: {PRIMARY} !important;
    text-align: center !important;
    letter-spacing: 0.20em !important;
    text-transform: uppercase !important;
    -webkit-text-stroke: 2.5px #000000;
    text-shadow:
        2px 2px 0 #000000,
        -2px 2px 0 #000000,
        2px -2px 0 #000000,
        -2px -2px 0 #000000;
    margin-top: -10px !important;
    margin-bottom: 12px !important;
}}

.app-subtitle {{
    font-size: 1.10rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.05em !important;
    color: #111827 !important;
    opacity: 0.97 !important;
    text-align: center !important;
    margin-bottom: 18px !important;
}}

[data-testid="stAppViewContainer"] {{
    background:
        linear-gradient(rgba(229,231,235,0.90), rgba(229,231,235,0.90)),
        url("data:image/jpeg;base64,{BG_B64}") no-repeat center fixed;
    background-size: 600px;
    color: #111827;
}}

.spray-card {{
    padding: 12px 14px;
    border-radius: 12px;
    border: 1px solid rgba(17,24,39,0.15);
    background: rgba(255,255,255,0.75);
}}

/* Make expander label bold */
[data-testid="stExpander"] summary {{
    font-weight: 800 !important;
}}
</style>

""",
    unsafe_allow_html=True,
)


# -----------------------------
# HEADER
# -----------------------------
st.markdown(f"<h1 class='app-title'>{SETTINGS.get('app_title','RP Spray Analytics')}</h1>", unsafe_allow_html=True)
st.markdown(f"<div class='app-subtitle'>{SETTINGS.get('subtitle','')}</div>", unsafe_allow_html=True)
st.markdown("---")

# -----------------------------
# SIDEBAR
# -----------------------------

import hashlib
import secrets
from datetime import datetime

# -----------------------------
# HALL OF FAME QUOTES (SIDEBAR)
# -----------------------------
HOF_QUOTES = [
    ("Hank Aaron", "Failure is a part of success."),
    ("Yogi Berra", "Baseball is 90% mental. The other half is physical."),
    ("Babe Ruth", "Never let the fear of striking out get in your way."),
    ("Ted Williams", "Hitting is timing. Pitching is upsetting timing."),
    ("Willie Mays", "It isn’t difficult to be great from time to time. What’s difficult is to be great all the time."),
    ("Cal Ripken Jr.", "Success is a process. You have to commit to the process."),
    ("Sandy Koufax", "Pitching is the art of instilling fear."),
    ("Nolan Ryan", "Enjoying success requires the ability to adapt."),
    ("Lou Gehrig", "It’s the ballplayer’s job to always be ready to play."),
    ("Jackie Robinson", "A life is not important except in the impact it has on other lives."),
]

def get_daily_quote(quotes):
    idx = int(datetime.utcnow().strftime("%Y%m%d")) % len(quotes)
    return quotes[idx]


# -----------------------------
# ACCESS CODE HASHING (ONE SOURCE OF TRUTH)
# -----------------------------
def hash_access_code(raw_code: str) -> str:
    salt = st.secrets.get("ACCESS_CODE_SALT", "")
    code = (raw_code or "").strip()
    if not salt:
        raise ValueError("Missing ACCESS_CODE_SALT in Streamlit secrets.")
    if not code:
        raise ValueError("Blank access code not allowed.")
    return hashlib.sha256((salt + "|" + code).encode("utf-8")).hexdigest()


def admin_set_access_code(team_slug: str, new_code: str) -> bool:
    team_slug = (team_slug or "").strip()
    if not team_slug:
        return False

    new_hash = hash_access_code(new_code)

    res = (
        supabase.table("team_access")
        .update({"code_hash": new_hash})
        .eq("team_slug", team_slug)
        .execute()
    )
    return bool(getattr(res, "data", None))


# -----------------------------
# SIDEBAR
# -----------------------------
import hashlib
import secrets
from datetime import datetime

# -----------------------------
# HALL OF FAME QUOTES (SIDEBAR)
# -----------------------------
HOF_QUOTES = [
    ("Hank Aaron", "Failure is a part of success."),
    ("Yogi Berra", "Baseball is 90% mental. The other half is physical."),
    ("Babe Ruth", "Never let the fear of striking out get in your way."),
    ("Ted Williams", "Hitting is timing. Pitching is upsetting timing."),
    ("Willie Mays", "It isn’t difficult to be great from time to time. What’s difficult is to be great all the time."),
    ("Cal Ripken Jr.", "Success is a process. You have to commit to the process."),
    ("Sandy Koufax", "Pitching is the art of instilling fear."),
    ("Nolan Ryan", "Enjoying success requires the ability to adapt."),
    ("Lou Gehrig", "It’s the ballplayer’s job to always be ready to play."),
    ("Jackie Robinson", "A life is not important except in the impact it has on other lives."),
]

def get_daily_quote(quotes):
    idx = int(datetime.utcnow().strftime("%Y%m%d")) % len(quotes)
    return quotes[idx]

# -----------------------------
# ACCESS CODE HASHING (ONE SOURCE OF TRUTH)
# -----------------------------
def hash_access_code(raw_code: str) -> str:
    salt = st.secrets.get("ACCESS_CODE_SALT", "")
    code = (raw_code or "").strip()
    if not salt:
        raise ValueError("Missing ACCESS_CODE_SALT in Streamlit secrets.")
    if not code:
        raise ValueError("Blank access code not allowed.")
    return hashlib.sha256((salt + "|" + code).encode("utf-8")).hexdigest()

def admin_set_access_code(team_slug: str, team_code: str, new_code: str) -> bool:
    """
    Updates team_access.code_hash for a team.
    Uses BOTH slug and code to hit the correct row no matter how the app identifies teams.
    """
    team_slug = (team_slug or "").strip()
    team_code = (team_code or "").strip().upper()

    if not team_slug and not team_code:
        return False

    new_hash = hash_access_code(new_code)

    q = supabase.table("team_access").update({"code_hash": new_hash})
    if team_slug:
        q = q.eq("team_slug", team_slug)
    if team_code:
        q = q.eq("team_code", team_code)

    res = q.execute()
    return bool(getattr(res, "data", None))

# -----------------------------
# SIDEBAR
# -----------------------------
import hashlib
import secrets
from datetime import datetime

# -----------------------------
# HALL OF FAME QUOTES (SIDEBAR)
# -----------------------------
HOF_QUOTES = [
    ("Hank Aaron", "Failure is a part of success."),
    ("Yogi Berra", "Baseball is 90% mental. The other half is physical."),
    ("Babe Ruth", "Never let the fear of striking out get in your way."),
    ("Ted Williams", "Hitting is timing. Pitching is upsetting timing."),
    ("Willie Mays", "It isn’t difficult to be great from time to time. What’s difficult is to be great all the time."),
    ("Cal Ripken Jr.", "Success is a process. You have to commit to the process."),
    ("Sandy Koufax", "Pitching is the art of instilling fear."),
    ("Nolan Ryan", "Enjoying success requires the ability to adapt."),
    ("Lou Gehrig", "It’s the ballplayer’s job to always be ready to play."),
    ("Jackie Robinson", "A life is not important except in the impact it has on other lives."),
]

def get_daily_quote(quotes):
    idx = int(datetime.utcnow().strftime("%Y%m%d")) % len(quotes)
    return quotes[idx]

# -----------------------------
# ACCESS CODE HASHING (ONE SOURCE OF TRUTH)
# -----------------------------
def hash_access_code(raw_code: str) -> str:
    salt = st.secrets.get("ACCESS_CODE_SALT", "")
    code = (raw_code or "").strip()
    if not salt:
        raise ValueError("Missing ACCESS_CODE_SALT in Streamlit secrets.")
    if not code:
        raise ValueError("Blank access code not allowed.")
    return hashlib.sha256((salt + "|" + code).encode("utf-8")).hexdigest()

def admin_set_access_code_by_id(row_id: int, new_code: str) -> bool:
    """Updates team_access.code_hash for a team by id (most reliable)."""
    try:
        rid = int(row_id)
    except Exception:
        return False

    new_hash = hash_access_code(new_code)
    res = supabase.table("team_access").update({"code_hash": new_hash}).eq("id", rid).execute()
    return bool(getattr(res, "data", None))


# -----------------------------
# SIDEBAR UI
# -----------------------------
with st.sidebar:
    # Logo
    if LOGO_PATH and os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=260)

    # Tags
    st.markdown("### ⚾ Spray Lab")
    st.markdown(
        """
        <span class="badge">Unlimited Teams</span>
        <span class="badge">GameChanger</span>
        <span class="badge">First Contact</span>
        <span class="badge">GB / FB</span>
        """,
        unsafe_allow_html=True,
    )

    # Strict mode
    strict_mode = st.checkbox(
        "STRICT MODE (only count plays with explicit fielder/location)",
        value=bool(SETTINGS.get("strict_mode_default", True)),
        key="strict_mode",
    )

    st.markdown("---")

    # Daily quote card
    who, quote = get_daily_quote(HOF_QUOTES)
    st.markdown(
        f"""
        <div style="
            padding: 14px;
            border-radius: 14px;
            background: rgba(255,255,255,0.72);
            border: 1px solid rgba(0,0,0,0.10);
            box-shadow: 0 6px 18px rgba(0,0,0,0.06);
        ">
            <div style="font-size: 0.95rem; font-weight: 800; margin-bottom: 8px;">
                🏆 Hall of Fame Quote
            </div>
            <div style="font-size: 0.98rem; font-weight: 700; line-height: 1.35;">
                “{quote}”
            </div>
            <div style="margin-top: 10px; font-size: 0.90rem; font-weight: 800; opacity: 0.85;">
                — {who}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

with st.sidebar:
    # ... your existing sidebar stuff above (logo, tags, quote, etc.)

    # -----------------------------
    # ADMIN SIDEBAR (BOTTOM)
    # -----------------------------
    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
    st.markdown("---")

    with st.expander("🔐 Admin", expanded=False):
        pin = st.text_input(
            "Admin PIN",
            type="password",
            label_visibility="collapsed",
            placeholder="Admin PIN",
            key="admin_pin_input",
        )

        if pin != st.secrets.get("ADMIN_PIN", ""):
            st.caption("Admin access only.")
        else:
            st.markdown(
                """
                <div style="
                    padding: 12px;
                    border-radius: 14px;
                    background: rgba(255,255,255,0.72);
                    border: 1px solid rgba(0,0,0,0.10);
                    box-shadow: 0 6px 18px rgba(0,0,0,0.06);
                    margin-bottom: 10px;
                ">
                    <div style="font-size:0.92rem; font-weight:800; margin-bottom:6px;">
                        Change Access Code
                    </div>
                    <div style="font-size:0.85rem; opacity:0.85;">
                        Updates Supabase instantly.
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # ✅ PERMANENT PARACHUTE (leave this)
            if st.button("🔄 Emergency Reset: Codes = TEAM CODE", key="admin_emergency_reset"):
                res = supabase.table("team_access").select("id, team_code").execute()
                rows = res.data or []

                updated = 0
                for r in rows:
                    rid = r.get("id")
                    code = (r.get("team_code") or "").strip().upper()
                    if rid and code:
                        supabase.table("team_access").update(
                            {"code_hash": hash_access_code(code)}
                        ).eq("id", rid).execute()
                        updated += 1

                load_team_codes.clear()
                st.success(f"Reset {updated} teams. Access code = TEAM CODE (ex: YUKON).")
                st.rerun()

            # Load teams DIRECT from Supabase
            res = (
                supabase.table("team_access")
                .select("id, team_code, team_name, is_active")
                .eq("is_active", True)
                .execute()
            )
            rows = res.data or []

            teams = []
            for r in rows:
                rid = r.get("id")
                code = (r.get("team_code") or "").strip().upper()
                name = (r.get("team_name") or "").strip()
                if rid and code:
                    label = f"{code} — {name}" if name else code
                    teams.append({"id": rid, "label": label})

            teams = sorted(teams, key=lambda x: x["label"])

            if not teams:
                st.error("No active teams found in team_access.")
            else:
                pick = st.selectbox(
                    "Team",
                    options=teams,
                    format_func=lambda x: x["label"],
                    key="admin_team_pick",
                )

                new_code = st.text_input("New Code", type="password", key="admin_new_code")
                confirm = st.text_input("Confirm", type="password", key="admin_confirm")

                c1, c2 = st.columns(2)
                update_btn = c1.button("💾 Update", use_container_width=True, key="admin_update_btn")
                clear_btn  = c2.button("Clear", use_container_width=True, key="admin_clear_btn")

                if clear_btn:
                    st.session_state["admin_new_code"] = ""
                    st.session_state["admin_confirm"] = ""
                    st.rerun()

                if update_btn:
                    if not (new_code or "").strip():
                        st.error("Enter a new code.")
                    elif new_code != confirm:
                        st.error("Codes don’t match.")
                    else:
                        ok = admin_set_access_code_by_id(pick["id"], new_code)
                        if ok:
                            st.success("✅ Access code updated.")
                            load_team_codes.clear()
                            st.rerun()
                        else:
                            st.error("Update failed.")


            st.markdown("### ➕ Add New School")
    
            with st.expander("Create School", expanded=False):
                colA, colB = st.columns(2)
                with colA:
                    new_team_name = st.text_input("School Name", key="new_team_name")
                    new_team_code = st.text_input("Team Code (ex: ROCK, YUKON)", key="new_team_code")
                with colB:
                    new_team_slug = st.text_input("Team Slug (unique)", key="new_team_slug")
                    new_active = st.checkbox("Active", value=True, key="new_team_active")
    
                new_logo = st.file_uploader("Team Logo", type=["png","jpg","jpeg","webp"], key="new_logo")
                new_bg   = st.file_uploader("Background Image", type=["png","jpg","jpeg","webp"], key="new_bg")
    
                if st.button("🚀 Create School", key="create_school_btn"):
                    if not (new_team_name or "").strip() or not (new_team_code or "").strip():
                        st.error("School name and team code are required.")
                    else:
                        team_slug = (new_team_slug or new_team_name.lower().replace(" ", "_")).strip()
                        team_code = new_team_code.upper().strip()
    
                        exists = (
                            supabase.table("team_access")
                            .select("id")
                            .eq("team_slug", team_slug)
                            .limit(1)
                            .execute()
                        )
                        if getattr(exists, "data", None):
                            st.error("That team slug already exists.")
                        else:
                            bucket = "team-assets"
                            try:
                                supabase.storage.create_bucket(bucket, public=True)
                            except Exception:
                                pass
    
                            logo_url = None
                            bg_url = None
    
                            if new_logo:
                                path = f"{team_slug}/logo.png"
                                supabase.storage.from_(bucket).upload(
                                    path,
                                    new_logo.getvalue(),
                                    file_options={"content-type": new_logo.type, "upsert": True},
                                )
                                logo_url = supabase.storage.from_(bucket).get_public_url(path)
    
                            if new_bg:
                                path = f"{team_slug}/background.png"
                                supabase.storage.from_(bucket).upload(
                                    path,
                                    new_bg.getvalue(),
                                    file_options={"content-type": new_bg.type, "upsert": True},
                                )
                                bg_url = supabase.storage.from_(bucket).get_public_url(path)
    
                            raw_key = secrets.token_hex(3).upper()
                            key_hash = hash_access_code(raw_key)
    
                            supabase.table("team_access").insert({
                                "team_slug": team_slug,
                                "team_code": team_code,
                                "team_name": new_team_name.strip(),
                                "code_hash": key_hash,
                                "is_active": bool(new_active),
                                "logo_url": logo_url,
                                "background_url": bg_url,
                            }).execute()
    
                            st.success("School created!")
                            st.code(f"Access Key: {raw_key}")
                            load_team_codes.clear()
                            st.rerun()


   
# -----------------------------
# TEAM SELECTION (SUPABASE - PERSISTENT)
# -----------------------------

teams = db_list_teams(TEAM_CODE_SAFE)

if not teams:
    st.warning("No teams found yet for THIS access code. Create one below.")
else:
    team_names = [t.get("team_name", "Unnamed Team") for t in teams]
    selected_team = st.selectbox("Choose a team:", team_names)

    selected_row = next((t for t in teams if t.get("team_name") == selected_team), teams[0])
    team_key = selected_row.get("team_key") or safe_team_key(selected_team)

with st.expander("➕ Add a new team (stored in Supabase)"):
    new_team_name = st.text_input("New team name:")
    if st.button("Create Team"):
        if not new_team_name.strip():
            st.error("Enter a team name first.")
        else:
            new_key = safe_team_key(new_team_name)
            db_upsert_team(TEAM_CODE_SAFE, new_key, new_team_name.strip(), "")
            st.success("Team created. Reloading…")
            st.rerun()

if not teams:
    st.stop()

st.markdown("---")

# -----------------------------
# ROSTER UI (SUPABASE - PERSISTENT)
# -----------------------------
st.subheader(f"📝 {selected_team} Roster (Hitters)")

default_roster_text = db_get_roster(TEAM_CODE_SAFE, team_key)

roster_text = st.text_area(
    "One player per line EXACTLY like GameChanger shows them (e.g., 'J Smith')",
    value=default_roster_text,
    height=220,
)

col_a, _ = st.columns([1, 3])
with col_a:
    if st.button("💾 Save Roster"):
        # Build the NEW roster from the text box (this is what coach just edited)
        new_roster = {ln.strip().strip('"') for ln in (roster_text or "").split("\n") if ln.strip()}

        # Save roster text
        db_upsert_team(TEAM_CODE_SAFE, team_key, selected_team, roster_text)

        # Reload season from DB (source of truth) – includes archived_players
        season_team, season_players, games_played, processed_set, archived_players = db_load_season_totals(
            TEAM_CODE_SAFE, team_key, new_roster
        )

        # Archive anyone removed from roster (but KEEP their stats)
        removed = set(season_players.keys()) - set(new_roster)
        archived_players = set(archived_players or set())
        archived_players.update(removed)

        # Unarchive anyone re-added
        archived_players = {p for p in archived_players if p not in new_roster}

        # Ensure new roster players exist in season_players
        for p in new_roster:
            season_players.setdefault(p, empty_stat_dict())

        # Save back with updated archived list
        db_save_season_totals(
            TEAM_CODE_SAFE, team_key, season_team, season_players, games_played, archived_players
        )

        st.success("Roster saved + removed players archived (reports will match roster).")
        st.rerun()

current_roster = {line.strip().strip('"') for line in roster_text.split("\n") if line.strip()}
st.write(f"**Hitters loaded:** {len(current_roster)}")


# ✅ LOAD FROM SUPABASE ONLY (source of truth) — includes archived_players
season_team, season_players, games_played, processed_set, archived_players = db_load_season_totals(
    TEAM_CODE_SAFE, team_key, current_roster
)

st.markdown(
    f"""
<div class="spray-card">
    <strong>Active team:</strong> {selected_team}<br>
    <strong>Games processed:</strong> {games_played}<br>
</div>
""",
    unsafe_allow_html=True,
)

# --- Reset button (Power Red, dynamic team name) ---
ACTIVE_TEAM_NAME = TEAM_CFG.get("team_name", TEAM_CODE)
reset_label = f"Reset SEASON totals"

st.markdown(
    """
    <style>
    button[aria-label^="Reset SEASON totals —"]{
        background-color:#b91c1c !important; /* Power Red */
        color:#ffffff !important;
        border:0 !important;
        font-weight:700 !important;
    }
    button[aria-label^="Reset SEASON totals —"]:hover{
        background-color:#991b1b !important;
        color:#ffffff !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

col_reset, _ = st.columns([1, 3])

with col_reset:
    if st.button("Reset Season Totals", key="reset_season", type="primary"):
        # Supabase is the source of truth now
        db_reset_season(TEAM_CODE_SAFE, team_key)

        season_team, season_players, games_played, processed_set, archived_players = db_load_season_totals(
            TEAM_CODE_SAFE, team_key, current_roster
        )

        st.rerun()


# -----------------------------
# PLAY-BY-PLAY INPUT
# -----------------------------
st.subheader("GameChanger Play-by-Play")

raw_text = st.text_area(
    f"Paste the full play-by-play for ONE game involving {selected_team}:",
    height=260,
)

# -----------------------------
# PROCESS GAME (SUPABASE DEDUPE + SUPABASE SAVE)
# -----------------------------
if "processing_game" not in st.session_state:
    st.session_state.processing_game = False
if "processing_started_at" not in st.session_state:
    st.session_state.processing_started_at = 0.0

# failsafe unlock after 15 seconds
if st.session_state.processing_game:
    try:
        if (time.time() - float(st.session_state.processing_started_at or 0.0)) > 15:
            st.session_state.processing_game = False
            st.session_state.processing_started_at = 0.0
    except Exception:
        st.session_state.processing_game = False
        st.session_state.processing_started_at = 0.0

st.markdown(
    """
    <style>
    #process-wrap button {
        background: #00c853 !important;
        color: white !important;
        border: 0 !important;
        font-weight: 700 !important;
        border-radius: 10px !important;
        padding: 0.6rem 1rem !important;
    }
    #process-wrap button:hover {
        background: #00b84a !important;
        color: white !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div id="process-wrap">', unsafe_allow_html=True)
process_clicked = st.button("📥 Process Game (ADD to Season Totals)", key="process_game_btn")
st.markdown("</div>", unsafe_allow_html=True)

if process_clicked:
    if st.session_state.processing_game:
        st.warning("Already processing… please wait.")
        st.stop()

    st.session_state.processing_game = True
    st.session_state.processing_started_at = time.time()

    rerun_needed = False
    marked_processed = False
    gkey = None

    try:
        if not (raw_text or "").strip():
            st.error("Paste play-by-play first.")
            st.stop()

        if not current_roster:
            st.error("Roster is empty. Add hitters first (and save).")
            st.stop()

        gkey = game_key_from_pbp(team_key, raw_text)

        if not db_try_mark_game_processed(TEAM_CODE_SAFE, team_key, gkey):
            st.warning("This exact play-by-play has already been processed for this team. Skipping.")
            st.stop()

        processed_set.add(gkey)
        marked_processed = True

        lines = [ln.strip() for ln in (raw_text or "").split("\n") if ln.strip()]

        game_team = empty_stat_dict()
        game_players = {p: empty_stat_dict() for p in current_roster}

        gp_in_game = set()
        running_seen = set()
        current_batter_ctx = None  # last known batter from "X at bat"

        for line in lines:
            clean_line = line.strip().strip('"')
            clean_line = re.sub(r"\([^)]*\)", "", clean_line)
            clean_line = re.sub(r"\s+", " ", clean_line).strip()
            if not clean_line:
                continue
            line_lower = clean_line.lower()
          
            # reset batter context at inning headers
            if line_lower.startswith("top ") or line_lower.startswith("bottom "):
                current_batter_ctx = None
                continue

            # --- GP tracking + batter context ---
            if not ("courtesy runner" in line_lower or re.search(r"\bcr\b", line_lower)):
                if " at bat" in line_lower:
                    bn = get_batter_name(clean_line, current_roster)
                    if bn:
                        gp_in_game.add(bn)
                        current_batter_ctx = bn

                if ("lineup changed" in line_lower) or ("defensive" in line_lower) or (" in for " in line_lower):
                    uline = (" " + clean_line.upper().replace(",", " ") + " ")
                    for p in current_roster:
                        if (" " + p.upper() + " ") in uline:
                            gp_in_game.add(p)

            # --- running events (NOT BIP) ---
            runner, total_key, base_key = parse_running_event(clean_line, current_roster)
            if runner and total_key:
                dedupe_key = (runner, total_key, base_key or "", line_lower)
                if dedupe_key not in running_seen:
                    running_seen.add(dedupe_key)

                    game_team[total_key] += 1
                    game_players[runner][total_key] += 1

                    if base_key and base_key in RUN_KEYS:
                        game_team[base_key] += 1
                        game_players[runner][base_key] += 1

            # --- resolve batter ---
            batter = get_batter_name(clean_line, current_roster) or current_batter_ctx
            if batter is None:
                continue

            gp_in_game.add(batter)

            if not is_ball_in_play(line_lower):
                continue

            # ✅ Bunts + Sac Bunts → ONE bucket
            if ("bunt" in line_lower) or ("sacrifice hit" in line_lower):
                game_team[BUNTS_KEY] += 1
                game_players[batter][BUNTS_KEY] += 1
                continue

            # --- normal GB/FB + location ---
            loc, loc_conf, _ = classify_location(line_lower, strict_mode=strict_mode)
            ball_type, bt_conf, _ = classify_ball_type(line_lower)

            if loc is None:
                if strict_mode:
                    continue
                loc = "UNKNOWN"

            if ball_type is None:
                if loc in ["SS", "3B", "2B", "1B", "P"]:
                    ball_type = "GB"
                elif loc in ["LF", "CF", "RF"]:
                    ball_type = "FB"

            game_team[loc] += 1
            game_players[batter][loc] += 1

            if ball_type in BALLTYPE_KEYS:
                game_team[ball_type] += 1
                game_players[batter][ball_type] += 1

            if ball_type in ("GB", "FB") and loc in COMBO_LOCS:
                ck = f"{ball_type}-{loc}"
                game_team[ck] += 1
                game_players[batter][ck] += 1

        # --- GP finalization ---
        for p in gp_in_game:
            if p in game_players:
                game_players[p][GP_KEY] += 1

        add_game_to_season(season_team, season_players, game_team, game_players)

        db_save_season_totals(
            TEAM_CODE_SAFE,
            team_key,
            season_team,
            season_players,
            len(processed_set),
            archived_players,
        )

        st.success("✅ Game processed and added to season totals.")
        rerun_needed = True

    except Exception as e:
        if marked_processed and gkey:
            processed_set.discard(gkey)
            db_unmark_game_processed(TEAM_CODE_SAFE, team_key, gkey)

        _show_db_error(e, "Processing failed (rolled back dedupe mark)")
        st.stop()

    finally:
        st.session_state.processing_game = False
        st.session_state.processing_started_at = 0.0

    if rerun_needed:
        st.cache_data.clear()
        st.rerun()



# -----------------------------
# SEASON OUTPUTS
# -----------------------------
hdr_left, hdr_right = st.columns([8, 2], vertical_alignment="center")
with hdr_left:
    st.markdown(
        f"<h3 style='margin:0; padding:0;'>📔 Full Team Spray – SEASON TO DATE ({selected_team})</h3>",
        unsafe_allow_html=True,
    )
with hdr_right:
    show_archived = st.checkbox("Show archived players", value=False)

# stat edit control placeholder (filled AFTER df_season exists)
stat_edit_slot = st.empty()

# -----------------------------
# Build season table (df_season)
# -----------------------------
season_rows = []

_roster_set = set(current_roster or [])
_season_players = season_players or {}
_archived_set = set(archived_players or set())

active_players = sorted([p for p in _roster_set if p in _season_players])
archived_list = sorted([p for p in _archived_set if p in _season_players and p not in _roster_set])

display_players = active_players + archived_list if show_archived else active_players

for player in display_players:
    stats = _season_players.get(player, {}) or {}
    row = {"Player": player}

    row["GB"] = stats.get("GB", 0)
    row["FB"] = stats.get("FB", 0)

    for ck in (COMBO_KEYS or []):
        row[ck] = stats.get(ck, 0)

    # ✅ ONE combined bunt stat (Bunt + Sac Bunt) + legacy fallbacks
    row["Bunts"] = (
        int(stats.get("Bunts", 0) or 0)
        + int(stats.get("BUNT", 0) or 0)
        + int(stats.get("Bunt", 0) or 0)
        + int(stats.get("Sac Bunt", 0) or 0)
        + int(stats.get("BU", 0) or 0)
        + int(stats.get("SH", 0) or 0)
    )

    for rk in (RUN_KEYS or []):
        row[rk] = stats.get(rk, 0)

    season_rows.append(row)

df_season = pd.DataFrame(season_rows)

# ✅ Use "Bunts" (not "BUNT") in the visible season table + exports
col_order = ["Player", "GB", "FB"] + list(COMBO_KEYS or []) + ["Bunts"] + list(RUN_KEYS or [])
if df_season.empty:
    df_season = pd.DataFrame(columns=col_order)
else:
    col_order = [c for c in col_order if c in df_season.columns]
    df_season = df_season[col_order]



# -----------------------------
# Stat Edit (column visibility) — NOW SAFE (df_season exists)
# -----------------------------
st.markdown(
    """
    <style>
      [data-testid="stDataFrameToolbar"] button[title="Download data as CSV"] { display: none !important; }
      [data-testid="stDataFrameToolbar"] button[aria-label="Download data as CSV"] { display: none !important; }
      [data-testid="stDataFrameToolbar"] button[title="Download data"] { display: none !important; }
      [data-testid="stDataFrameToolbar"] button[aria-label="Download data"] { display: none !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <style>
    .stat-edit-wrap {
        display: flex;
        justify-content: flex-end;
        align-items: center;
        margin-top: 0px !important;
        margin-bottom: 0px !important;
    }
    .stat-edit-wrap button {
        white-space: nowrap;
        border-radius: 999px !important;
        padding: 0.35rem 0.75rem !important;
        font-weight: 800 !important;
        font-size: 0.75rem !important;
        letter-spacing: 0.08em !important;
        text-transform: uppercase !important;
        background: rgba(17,24,39,0.06) !important;
        border: 1px solid rgba(17,24,39,0.18) !important;
        color: rgba(17,24,39,0.92) !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
    }
    .stat-edit-wrap button:hover {
        background: rgba(17,24,39,0.10) !important;
        border-color: rgba(17,24,39,0.28) !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

cols_key = f"season_cols__{TEAM_CODE_SAFE}__{team_key}"
all_cols = list(df_season.columns)

# ✅ Auto-add any NEW columns to the saved Stat Edit selection (so new stats show up)
_saved = st.session_state.get(cols_key, [])
if isinstance(_saved, (list, tuple)):
    missing = [c for c in all_cols if c not in _saved]
    if missing:
        st.session_state[cols_key] = list(_saved) + missing
else:
    st.session_state[cols_key] = all_cols.copy()

if cols_key not in st.session_state:
    st.session_state[cols_key] = all_cols.copy()

default_cols = list(st.session_state.get(cols_key, []))
default_cols = [c for c in default_cols if c in all_cols]

if "Player" in all_cols and "Player" not in default_cols:
    default_cols = ["Player"] + default_cols




# -----------------------------
# STAT FILTERS (Popover / Expander)
# -----------------------------
with stat_edit_slot.container():
    if hasattr(st, "popover"):
        with st.popover("⚙ Stat Filters"):
            st.caption("Toggle which stats show in the table")
            flt = st.text_input(
                "Search",
                value="",
                placeholder="Type to filter stats...",
                key=f"{cols_key}__flt",
            )

            c1, c2 = st.columns(2)
            with c1:
                all_clicked = st.button("All", key=f"{cols_key}__all", use_container_width=True)
            with c2:
                none_clicked = st.button("None", key=f"{cols_key}__none", use_container_width=True)

            if all_clicked or none_clicked:
                for _col in all_cols:
                    _safe = re.sub(r"[^A-Za-z0-9_]+", "_", str(_col))
                    _k = f"{cols_key}__cb__{_safe}"
                    st.session_state[_k] = True if (_col == "Player" or all_clicked) else False

                st.session_state[cols_key] = list(all_cols) if all_clicked else (["Player"] if "Player" in all_cols else [])
                st.rerun()

            picked_set = set(st.session_state.get(cols_key, default_cols))
            if "Player" in all_cols:
                picked_set.add("Player")

            view_cols = list(all_cols)
            if flt.strip():
                q = flt.strip().lower()
                view_cols = [c for c in view_cols if q in str(c).lower()]

            with st.container(height=360):
                if "Player" in view_cols:
                    st.checkbox("Player", value=True, disabled=True, key=f"{cols_key}__cb__Player")
                    view_cols = [c for c in view_cols if c != "Player"]

                colA, colB, colC = st.columns(3)
                grid = [colA, colB, colC]

                for i, col in enumerate(view_cols):
                    target = grid[i % 3]
                    safe_col = re.sub(r"[^A-Za-z0-9_]+", "_", str(col))
                    cur_val = col in picked_set
                    new_val = target.checkbox(
                        str(col),
                        value=cur_val,
                        key=f"{cols_key}__cb__{safe_col}",
                    )
                    if new_val:
                        picked_set.add(col)
                    else:
                        picked_set.discard(col)

            st.session_state[cols_key] = [c for c in all_cols if c in picked_set]

    else:
        with st.expander("⚙ Stat Filters", expanded=False):
            st.caption("Toggle which stats show in the table")
            flt = st.text_input(
                "Search",
                value="",
                placeholder="Type to filter stats...",
                key=f"{cols_key}__flt",
            )

            c1, c2 = st.columns(2)
            with c1:
                all_clicked = st.button("All", key=f"{cols_key}__all", use_container_width=True)
            with c2:
                none_clicked = st.button("None", key=f"{cols_key}__none", use_container_width=True)

            if all_clicked or none_clicked:
                for _col in all_cols:
                    _safe = re.sub(r"[^A-Za-z0-9_]+", "_", str(_col))
                    _k = f"{cols_key}__cb__{_safe}"
                    st.session_state[_k] = True if (_col == "Player" or all_clicked) else False

                st.session_state[cols_key] = list(all_cols) if all_clicked else (["Player"] if "Player" in all_cols else [])
                st.rerun()

            picked_set = set(st.session_state.get(cols_key, default_cols))
            if "Player" in all_cols:
                picked_set.add("Player")

            view_cols = list(all_cols)
            if flt.strip():
                q = flt.strip().lower()
                view_cols = [c for c in view_cols if q in str(c).lower()]

            with st.container(height=360):
                if "Player" in view_cols:
                    st.checkbox("Player", value=True, disabled=True, key=f"{cols_key}__cb__Player")
                    view_cols = [c for c in view_cols if c != "Player"]

                colA, colB, colC = st.columns(3)
                grid = [colA, colB, colC]

                for i, col in enumerate(view_cols):
                    target = grid[i % 3]
                    safe_col = re.sub(r"[^A-Za-z0-9_]+", "_", str(col))
                    cur_val = col in picked_set
                    new_val = target.checkbox(
                        str(col),
                        value=cur_val,
                        key=f"{cols_key}__cb__{safe_col}",
                    )
                    if new_val:
                        picked_set.add(col)
                    else:
                        picked_set.discard(col)

            st.session_state[cols_key] = [c for c in all_cols if c in picked_set]


# -----------------------------
# APPLY COLUMN SELECTION
# -----------------------------
picked_cols = [
    c for c in st.session_state.get(cols_key, []) if c in df_season.columns
]

if "Player" in df_season.columns and "Player" not in picked_cols:
    picked_cols = ["Player"] + picked_cols

df_show = df_season[picked_cols] if picked_cols else df_season

# -----------------------------
# VISIBLE COLS (for CSV / downloads)
# -----------------------------
if df_show is not None and not df_show.empty:
    visible_cols = list(df_show.columns)
else:
    visible_cols = list(df_season.columns) if df_season is not None else []

# -----------------------------
# TABLE RENDER (NO EMPTY GAP)
# -----------------------------
if df_show is None or df_show.empty:
    st.info("No season stats to display yet. Process at least one game to generate season totals.")
else:
    st.dataframe(df_show, use_container_width=True)


# -----------------------------
# 📝 COACHES SCOUTING NOTES (per selected opponent/team)
# -----------------------------
notes_key = f"coaches_notes__{TEAM_CODE_SAFE}__{team_key}"
if notes_key not in st.session_state:
    st.session_state[notes_key] = db_get_coach_notes(TEAM_CODE_SAFE, team_key)

player_notes_key = f"player_notes__{TEAM_CODE_SAFE}__{team_key}"
if player_notes_key not in st.session_state:
    st.session_state[player_notes_key] = db_get_player_notes(TEAM_CODE_SAFE, team_key)

with st.expander("📝 Coaches Scouting Notes (prints on Excel/CSV)", expanded=False):
    st.session_state[notes_key] = st.text_area(
        "Notes for THIS selected opponent/team:",
        value=st.session_state[notes_key],
        height=160,
        key=f"{notes_key}__box",
    )

    if st.button("💾 Save Notes", key=f"{notes_key}__save"):
        db_save_season_totals(
            TEAM_CODE_SAFE,
            team_key,
            season_team,
            season_players,
            games_played,
            archived_players,
            coach_notes=st.session_state[notes_key],
        )
        st.success("Notes saved for this opponent/team.")

notes_box_text = str(st.session_state.get(notes_key, "") or "").strip()

_csv_text = (
    df_season[[c for c in visible_cols if c in df_season.columns]].to_csv(index=False)
    if (df_season is not None and not df_season.empty)
    else ""
)

# CSV can't merge cells, but we can push notes to the bottom for printing
if notes_box_text:
    import csv as _csv
    import io as _io
    cols = list(df_season.columns)
    blank_row = [""] * len(cols)

    # Build a footer row: COACH NOTES + note text
    if len(cols) == 0:
        cols = ["Player"]
    footer = [""] * len(cols)
    if len(cols) == 1:
        footer[0] = "COACH NOTES: " + notes_box_text.replace("\n", " ")
    else:
        footer[0] = "COACH NOTES:"
        footer[1] = notes_box_text.replace("\n", "  ")

    buf = _io.StringIO()
    w = _csv.writer(buf, lineterminator="\n")
    w.writerow([])  # ensure we start on new line cleanly
    for _ in range(5):
        w.writerow(blank_row)
    w.writerow(footer)

    _csv_text = _csv_text.rstrip("\n") + "\n" + buf.getvalue().lstrip("\n")

csv_bytes = _csv_text.encode("utf-8")
safe_team = re.sub(r"[^A-Za-z0-9_-]+", "_", selected_team).strip("_")

# --- Download should match current Stat Edit view ---
# Build a safe visible_cols list (prevents NameError and handles empty seasons cleanly)
try:
    _vc = st.session_state.get(cols_key, list(df_season.columns))
except Exception:
    _vc = list(df_season.columns)

if not isinstance(_vc, (list, tuple)):
    _vc = list(df_season.columns)

visible_cols = [c for c in _vc if c in df_season.columns]

# Always keep Player if it exists
if "Player" in df_season.columns and "Player" not in visible_cols:
    visible_cols = ["Player"] + visible_cols

no_season_data = (df_season is None) or (getattr(df_season, "empty", True)) or (len(getattr(df_season, "columns", [])) == 0)

if no_season_data:
    st.info("No season stats to download yet. Process at least one game to generate season totals.")
    # Fallback so the app doesn't crash — still allows the page to load.
    df_xl = df_season.copy() if df_season is not None else None
else:
    df_xl = df_season[visible_cols].copy()


out = BytesIO()

# -----------------------------
# SEASON REPORT (EXCEL) — PRINT-STYLE FORMATTING
# -----------------------------
def _safe_sheet_name(name: str, used: set[str]) -> str:
    # Excel: max 31 chars, no : \ / ? * [ ]
    base = re.sub(r'[:\\/*?\[\]]', '', str(name or "").strip())
    if not base:
        base = "Player"
    base = base[:31]
    nm = base
    k = 2
    while nm in used:
        suffix = f"_{k}"
        nm = (base[: 31 - len(suffix)] + suffix)[:31]
        k += 1
    used.add(nm)
    return nm


def _build_individual_spray_sheet(
    wb,
    sheet_name,
    player_name,
    stats,
    notes_text="",
    template_mode=False  # ✅ ADD THIS
):


    """
    Builds the EXACT style 'Individual Spray' tab:
    - Header bar with player name
    - Position boxes with GB/FB counts and % of total BIP
    - BIP total box
    - Event log grid + Notes box
    - Heatmap on % cells
    """

    ws = wb.create_sheet(title=sheet_name)

    if template_mode:
        stats = {}
        player_name = ""


    # -----------------------------
    # Layout constants (matches your screenshot grid style)
    # -----------------------------
    COL_LEFT = 2   # B
    COL_RIGHT = 11  # k

    # Header bar
    HEADER_TOP = 2
    HEADER_BOT = 3

    # Position boxes (top-left anchored like screenshot)
    pos_boxes = {
        "LF": (4, 4),   # D-E
        "CF": (4, 6),   # F-G   ✅ moved so it doesn't overlap LF
        "RF": (4, 8),   # H-I

        "SS": (8, 5),   # E-F
        "2B": (8, 7),   # G-H

        "3B": (12, 4),  # D-E
        "P":  (14, 6),  # F-G
        "1B": (12, 8),  # H-I
    }

    # BIP Total box
    BIP_ROW = 17
    BIP_COL = 2  # B

    # Log table area
    LOG_TOP = 21
    LOG_LEFT = 2   # B
    LOG_RIGHT = 10  # J
    LOG_ROWS = 20

    # -----------------------------
    # Styling
    # -----------------------------
    thin = Side(style="thin", color="000000")
    thick = Side(style="thick", color="000000")

    def border_box(r1, c1, r2, c2, thick_outer=True):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = Border(
                    left=(thick if thick_outer and c == c1 else thin),
                    right=(thick if thick_outer and c == c2 else thin),
                    top=(thick if thick_outer and r == r1 else thin),
                    bottom=(thick if thick_outer and r == r2 else thin),
                )

    # Column widths
    ws.column_dimensions["A"].width = 2.5
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 10

    # -----------------------------
    # Header bar  ✅ value BEFORE merge
    # -----------------------------
    hc = ws.cell(row=HEADER_TOP, column=COL_LEFT)
    hc.value = str(player_name)
    hc.font = Font(bold=True, size=20)
    hc.alignment = Alignment(horizontal="center", vertical="center")
    hc.fill = PatternFill("solid", fgColor="D9D9D9")

    ws.merge_cells(
    start_row=HEADER_TOP, start_column=COL_LEFT,
    end_row=HEADER_BOT, end_column=COL_RIGHT
    )
    border_box(HEADER_TOP, COL_LEFT, HEADER_BOT, COL_RIGHT, thick_outer=True)



    # -----------------------------
    # FINAL SPEC — Individual tab Excel adjustments
    # -----------------------------

    # 1️⃣ Header Row 1: merge A1:J1 (rows 2–3 header stays as-is)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=1, column=1, value="").alignment = Alignment(horizontal="center", vertical="center")

    # 2️⃣ Row heights = 24 for rows 5, 9, 13, 15
    for rr in [5, 9, 13, 15]:
        ws.row_dimensions[rr].height = 24

    # ✅ Result MUST be in K20
    ws.column_dimensions["K"].width = 12
    rcell = ws.cell(row=20, column=11, value="Result")  # K20
    rcell.font = Font(bold=True, size=10)
    rcell.alignment = Alignment(horizontal="center", vertical="center")

    # 3️⃣ Bottom merges + numbering (keep EXACTLY what you see: 21–40)
    merge_pairs = [(21, 22), (23, 24), (25, 26), (27, 28), (29, 30),
                   (31, 32), (33, 34), (35, 36), (37, 38), (39, 40)]

    for i, (top, bot) in enumerate(merge_pairs, start=1):
        # Merge Column B
        ws.merge_cells(start_row=top, start_column=2, end_row=bot, end_column=2)  # B
        ncell = ws.cell(row=top, column=2, value=i)
        ncell.font = Font(size=12)
        ncell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # ✅ Merge Column K (same pairs)
        ws.merge_cells(start_row=top, start_column=11, end_row=bot, end_column=11)  # K

        # 4️⃣ Column C labels: B on top row, S on bottom row (KEEP S)
        bcell = ws.cell(row=top, column=3, value="B")
        bcell.font = Font(bold=True, size=10)
        bcell.alignment = Alignment(horizontal="center", vertical="center")

        scell = ws.cell(row=bot, column=3, value="S")
        scell.font = Font(bold=True, size=10)
        scell.alignment = Alignment(horizontal="center", vertical="center")

    # -----------------------------
    # Percent heatmap bins (same as Season style)
    # -----------------------------
    pct_bins = [
        (0.00, 0.05, None),
        (0.05, 0.10, PatternFill("solid", fgColor="FFE5CC")),
        (0.10, 0.15, PatternFill("solid", fgColor="FFDBB8")),
        (0.15, 0.20, PatternFill("solid", fgColor="FFCC99")),
        (0.20, 0.25, PatternFill("solid", fgColor="FFBE80")),
        (0.25, 0.30, PatternFill("solid", fgColor="FFB266")),
        (0.30, 0.35, PatternFill("solid", fgColor="FFA366")),
        (0.35, 0.40, PatternFill("solid", fgColor="FF9933")),
        (0.40, 0.45, PatternFill("solid", fgColor="F8A5A5")),
        (0.45, 0.50, PatternFill("solid", fgColor="F28B82")),
        (0.50, 0.55, PatternFill("solid", fgColor="F8696B")),
        (0.55, 0.60, PatternFill("solid", fgColor="EF5350")),
        (0.60, 0.65, PatternFill("solid", fgColor="E53935")),
        (0.65, 0.70, PatternFill("solid", fgColor="D32F2F")),
        (0.70, 0.75, PatternFill("solid", fgColor="C62828")),
        (0.75, 0.80, PatternFill("solid", fgColor="B71C1C")),
        (0.80, 0.85, PatternFill("solid", fgColor="A00000")),
        (0.85, 0.90, PatternFill("solid", fgColor="8E0000")),
        (0.90, 0.95, PatternFill("solid", fgColor="7F0000")),
        (0.95, 1.00, PatternFill("solid", fgColor="6A0000")),
    ]

    def pct_fill(v):
        try:
            x = float(v)
        except Exception:
            return None
        if x <= 0:
            return None
        if x < 0:
            x = 0.0
        if x > 1:
            x = 1.0
        for lo, hi, fill in pct_bins:
            if fill is None:
                continue
            if (lo <= x < hi) or (hi == 1.0 and lo <= x <= hi):
                return fill
        return None

    # -----------------------------
    # Totals
    # -----------------------------
    gb_total = int(stats.get("GB", 0) or 0)
    fb_total = int(stats.get("FB", 0) or 0)
    bip_total = gb_total + fb_total
    denom = bip_total if bip_total > 0 else 0

    # -----------------------------
    # Position boxes
    # -----------------------------
    title_font = Font(bold=True, size=10)
    small_font = Font(bold=True, size=9)
    val_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for pos, (r0, c0) in pos_boxes.items():
        r1, c1 = r0, c0
        r2, c2 = r0 + 2, c0 + 1

        # ✅ value BEFORE merge (prevents MergedCell crash)
        tcell = ws.cell(row=r1, column=c1)
        tcell.value = pos
        tcell.font = title_font
        tcell.alignment = center
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r1, end_column=c2)

        gb_k = f"GB-{pos}"
        fb_k = f"FB-{pos}"
        gb_ct = int(stats.get(gb_k, 0) or 0)
        fb_ct = int(stats.get(fb_k, 0) or 0)

        # Counts row
        cL = ws.cell(row=r1 + 1, column=c1)
        cL.value = f"GB\n{gb_ct}"
        cL.font = small_font
        cL.alignment = center

        cR = ws.cell(row=r1 + 1, column=c2)
        cR.value = f"FB\n{fb_ct}"
        cR.font = small_font
        cR.alignment = center

        # % row of TOTAL BIP
        gb_pct = (gb_ct / denom) if denom else 0.0
        fb_pct = (fb_ct / denom) if denom else 0.0

        left_pct_cell = ws.cell(row=r1 + 2, column=c1)
        left_pct_cell.value = gb_pct
        left_pct_cell.number_format = "0%"
        left_pct_cell.font = val_font
        left_pct_cell.alignment = center

        right_pct_cell = ws.cell(row=r1 + 2, column=c2)
        right_pct_cell.value = fb_pct
        right_pct_cell.number_format = "0%"
        right_pct_cell.font = val_font
        right_pct_cell.alignment = center

        f1 = pct_fill(gb_pct)
        f2 = pct_fill(fb_pct)
        if f1:
            left_pct_cell.fill = f1
        if f2:
            right_pct_cell.fill = f2

        border_box(r1, c1, r2, c2, thick_outer=True)

    # -----------------------------
    # BIP Total box ✅ value BEFORE merge
    # -----------------------------
    lab = ws.cell(row=BIP_ROW, column=BIP_COL)
    lab.value = "BIP - Total"
    lab.font = Font(bold=True, size=10)
    lab.alignment = center
    ws.merge_cells(
        start_row=BIP_ROW, start_column=BIP_COL,
        end_row=BIP_ROW, end_column=BIP_COL + 1
    )

    val = ws.cell(row=BIP_ROW + 1, column=BIP_COL)
    val.value = int(bip_total)
    val.font = Font(bold=True, size=12)
    val.alignment = center
    ws.merge_cells(
        start_row=BIP_ROW + 1, start_column=BIP_COL,
        end_row=BIP_ROW + 1, end_column=BIP_COL + 1
    )

    border_box(BIP_ROW, BIP_COL, BIP_ROW + 1, BIP_COL + 1, thick_outer=True)

    # -----------------------------
    # SB / CS totals (Row 17-18, Col I-J)
    # -----------------------------
    sb_total = int(stats.get("SB", 0) or 0)
    cs_total = int(stats.get("CS", 0) or 0)

    # Labels
    sb_lab = ws.cell(row=17, column=9, value="SB")   # I17
    cs_lab = ws.cell(row=17, column=10, value="CS")  # J17
    for cell in (sb_lab, cs_lab):
        cell.font = Font(bold=True, size=10)
        cell.alignment = center
        cell.fill = PatternFill("solid", fgColor="D9D9D9")  # matches your gray style
        cell.border = Border(left=thick, right=thick, top=thick, bottom=thin)

    # Values
    sb_val = ws.cell(row=18, column=9, value=sb_total)   # I18
    cs_val = ws.cell(row=18, column=10, value=cs_total)  # J18
    for cell in (sb_val, cs_val):
        cell.font = Font(bold=True, size=12)
        cell.alignment = center
        cell.border = Border(left=thick, right=thick, top=thin, bottom=thick)

    # -----------------------------
    # Bunts total (Bunt + Sac Bunt combined) — Row 17-18, Col K
    # -----------------------------
    bun_total = (
        int(stats.get("Bunts", 0) or 0)
        + int(stats.get("BUNT", 0) or 0)
        + int(stats.get("Bunt", 0) or 0)
        + int(stats.get("Sac Bunt", 0) or 0)
        + int(stats.get("BU", 0) or 0)
        + int(stats.get("SH", 0) or 0)
    )
    
    # Label (K17)
    bun_lab = ws.cell(row=17, column=11, value="BUNTS")   # K17
    bun_lab.font = Font(bold=True, size=10)
    bun_lab.alignment = center
    bun_lab.fill = PatternFill("solid", fgColor="D9D9D9")
    bun_lab.border = Border(left=thick, right=thick, top=thick, bottom=thin)
    
    # Value (K18)
    bun_val = ws.cell(row=18, column=11, value=bun_total)  # K18
    bun_val.font = Font(bold=True, size=12)
    bun_val.alignment = center
    bun_val.border = Border(left=thick, right=thick, top=thin, bottom=thick)
    
    # ✅ If you previously used H17/H18 for SAC, clear them so nothing shows
    ws.cell(row=17, column=8, value="")
    ws.cell(row=18, column=8, value="")
        
    # -----------------------------
    # Black divider bar
    # -----------------------------
    bar = ws.cell(row=LOG_TOP - 1, column=LOG_LEFT)
    bar.value = ""
    bar.fill = PatternFill("solid", fgColor="000000")
    ws.merge_cells(
        start_row=LOG_TOP - 1, start_column=LOG_LEFT,
        end_row=LOG_TOP - 1, end_column=LOG_RIGHT
    )

    # -----------------------------
    # Log grid
    # -----------------------------
    # NOTE: Result header is handled above (K20). Do NOT write it again here.

    for r in range(LOG_TOP, LOG_TOP + LOG_ROWS):
        for c in range(LOG_LEFT, LOG_RIGHT + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.font = Font(size=10)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    border_box(LOG_TOP, LOG_LEFT, LOG_TOP + LOG_ROWS - 1, LOG_RIGHT, thick_outer=True)

    res_col = LOG_RIGHT + 1  # K (because LOG_RIGHT=J=10)
    ws.column_dimensions[get_column_letter(res_col)].width = 12
    for r in range(LOG_TOP, LOG_TOP + LOG_ROWS):
        cell = ws.cell(row=r, column=res_col)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = center
        cell.font = Font(size=10)
    border_box(LOG_TOP, res_col, LOG_TOP + LOG_ROWS - 1, res_col, thick_outer=True)

    # -----------------------------
    # Print setup (portrait, fit)
    # -----------------------------
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER

    return ws


out = BytesIO()

with pd.ExcelWriter(out, engine="openpyxl") as writer:
    sheet_name = "Season"

    # Build export frame
    df_export = df_xl.copy() if df_xl is not None else pd.DataFrame()

    # Insert GP (Games Played) after Player
    if not df_export.empty and "Player" in df_export.columns:
        def _gp_for(name):
            try:
                return int((season_players.get(str(name), {}) or {}).get(GP_KEY, 0) or 0)
            except Exception:
                return 0
        # avoid duplicate GP insert if rerun logic ever touches this again
        if "GP" not in df_export.columns:
            df_export.insert(1, "GP", df_export["Player"].apply(_gp_for))

    # --- Build BIP + GB%/FB% (based on total BIP = GB + FB) ---
    if not df_export.empty and ("GB" in df_export.columns) and ("FB" in df_export.columns):
        gb_vals = pd.to_numeric(df_export["GB"], errors="coerce").fillna(0)
        fb_vals = pd.to_numeric(df_export["FB"], errors="coerce").fillna(0)

        bip_vals = (gb_vals + fb_vals).fillna(0)
        denom = bip_vals.replace({0: pd.NA})

        # Percent columns
        df_export["GB%"] = (gb_vals / denom).fillna(0)
        df_export["FB%"] = (fb_vals / denom).fillna(0)

        # Convert positional columns (GB-* and FB-*) to % of TOTAL BIP
        for c in list(df_export.columns):
            if str(c).startswith("GB-") or str(c).startswith("FB-"):
                num = pd.to_numeric(df_export[c], errors="coerce").fillna(0)
                df_export[c] = (num / denom).fillna(0)

        # Drop raw GB/FB totals
        df_export = df_export.drop(columns=["GB", "FB"])

        # Put columns in desired order
        cols = list(df_export.columns)
        gb_pos = [c for c in cols if str(c).startswith("GB-")]
        fb_pos = [c for c in cols if str(c).startswith("FB-")]

        fixed_lead = ["Player"] + (["GP"] if "GP" in cols else []) + ["GB%", "FB%"]
        rest = [c for c in cols if c not in fixed_lead and c not in gb_pos and c not in fb_pos]

        # Add BIP at the end of FB block
        df_export["BIP"] = bip_vals.astype(int)

        # ✅ Put BUNT + SB/CS immediately to the right of BIP
        bunt_and_run = ["BUNT", "SB", "SB-2B", "SB-3B", "CS", "CS-2B", "CS-3B"]
        rest2 = [c for c in rest if c not in bunt_and_run and c != "BIP"]
        present_br = [c for c in bunt_and_run if c in df_export.columns]

        df_export = df_export[fixed_lead + gb_pos + fb_pos + ["BIP"] + present_br + rest2]

    # Write Season sheet
    df_export.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
    ws = writer.book[sheet_name]

    # Safety: ensure visible
    try:
        for _sh in writer.book.worksheets:
            _sh.sheet_state = "visible"
        writer.book.active = writer.book.worksheets.index(ws)
    except Exception:
        pass

    # Team title row (Row 1)
    total_cols = max(1, ws.max_column)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=str(selected_team))
    title_cell.font = Font(bold=True, size=28)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"

    # Row heights
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 35
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 45  # ✅ Player rows height

    # Header styling (Row 2)
    header_font = Font(bold=True, size=12)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill

    # Player column formatting
    player_col_idx = None
    for j in range(1, ws.max_column + 1):
        if str(ws.cell(row=2, column=j).value).strip() == "Player":
            player_col_idx = j
            break

    body_font = Font(size=12)
    player_font = Font(size=12, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.alignment = center_align
            if player_col_idx and c == player_col_idx:
                cell.font = player_font
                cell.alignment = left_align

    # Autosize Player col
    if player_col_idx:
        max_len = len("Player")
        try:
            series = df_export["Player"].astype(str).tolist() if "Player" in df_export.columns else []
            for v in series[:200]:
                max_len = max(max_len, len(v))
        except Exception:
            pass
        ws.column_dimensions[get_column_letter(player_col_idx)].width = min(max(max_len + 2, 12), 34)

    # Identify key columns
    gbp_idx = None
    fbp_idx = None
    gp_idx = None
    bip_idx = None

    headers = [str(ws.cell(row=2, column=j).value or "").strip() for j in range(1, ws.max_column + 1)]
    for j, h in enumerate(headers, start=1):
        if h == "GB%":
            gbp_idx = j
        elif h == "FB%":
            fbp_idx = j
        elif h == "GP":
            gp_idx = j
        elif h == "BIP":
            bip_idx = j

    # Format GB%/FB% as percent
    if gbp_idx:
        L = get_column_letter(gbp_idx)
        for r in range(3, ws.max_row + 1):
            ws[f"{L}{r}"].number_format = "0%"
    if fbp_idx:
        L = get_column_letter(fbp_idx)
        for r in range(3, ws.max_row + 1):
            ws[f"{L}{r}"].number_format = "0%"

    # Format positional % columns as percent too
    for j, h in enumerate(headers, start=1):
        if h.startswith("GB-") or h.startswith("FB-"):
            L = get_column_letter(j)
            for r in range(3, ws.max_row + 1):
                ws[f"{L}{r}"].number_format = "0%"

    # -----------------------------
    # ✅ BORDERS (all INSIDE the with-block, so no indentation errors)
    # -----------------------------
    thick_side = Side(style="thick", color="000000")

    def _outline_box(r1: int, c1: int, r2: int, c2: int):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=rr, column=cc)
                b = cell.border
                cell.border = Border(
                    left=thick_side if cc == c1 else b.left,
                    right=thick_side if cc == c2 else b.right,
                    top=thick_side if rr == r1 else b.top,
                    bottom=thick_side if rr == r2 else b.bottom,
                )

    def _first_idx(prefix: str):
        for jj, hh in enumerate(headers, start=1):
            if hh.startswith(prefix):
                return jj
        return None

    def _last_idx(prefix: str):
        last = None
        for jj, hh in enumerate(headers, start=1):
            if hh.startswith(prefix):
                last = jj
        return last

    def _set_right_thick(col_idx: int):
        for rr in range(2, ws.max_row + 1):  # include header row
            cell = ws.cell(row=rr, column=col_idx)
            b = cell.border
            cell.border = Border(left=b.left, right=thick_side, top=b.top, bottom=b.bottom)

    gb_start = _first_idx("GB-")
    gb_end = _last_idx("GB-")
    fb_start = _first_idx("FB-")
    fb_end = _last_idx("FB-")

    # Thick line after BIP to separate BIP and SB/CS
    if bip_idx:
        _set_right_thick(bip_idx)

    # Thick outline around GB and FB blocks (including headings row 2)
    if gb_start and gb_end:
        _outline_box(2, gb_start, ws.max_row, gb_end)
    if fb_start and fb_end:
        _outline_box(2, fb_start, ws.max_row, fb_end)

    # -----------------------------
    # HEATMAPS
    # -----------------------------
    gp_fill_1_5   = PatternFill("solid", fgColor="FFE5CC")
    gp_fill_6_10  = PatternFill("solid", fgColor="FFCC99")
    gp_fill_11_15 = PatternFill("solid", fgColor="FFB266")
    gp_fill_16_19 = PatternFill("solid", fgColor="FF9933")
    gp_fill_20p   = PatternFill("solid", fgColor="F8696B")

    pct_bins = [
        (0.00, 0.05, None),
        (0.05, 0.10, PatternFill("solid", fgColor="FFE5CC")),
        (0.10, 0.15, PatternFill("solid", fgColor="FFDBB8")),
        (0.15, 0.20, PatternFill("solid", fgColor="FFCC99")),
        (0.20, 0.25, PatternFill("solid", fgColor="FFBE80")),
        (0.25, 0.30, PatternFill("solid", fgColor="FFB266")),
        (0.30, 0.35, PatternFill("solid", fgColor="FFA366")),
        (0.35, 0.40, PatternFill("solid", fgColor="FF9933")),
        (0.40, 0.45, PatternFill("solid", fgColor="F8A5A5")),
        (0.45, 0.50, PatternFill("solid", fgColor="F28B82")),
        (0.50, 0.55, PatternFill("solid", fgColor="F8696B")),
        (0.55, 0.60, PatternFill("solid", fgColor="EF5350")),
        (0.60, 0.65, PatternFill("solid", fgColor="E53935")),
        (0.65, 0.70, PatternFill("solid", fgColor="D32F2F")),
        (0.70, 0.75, PatternFill("solid", fgColor="C62828")),
        (0.75, 0.80, PatternFill("solid", fgColor="B71C1C")),
        (0.80, 0.85, PatternFill("solid", fgColor="A00000")),
        (0.85, 0.90, PatternFill("solid", fgColor="8E0000")),
        (0.90, 0.95, PatternFill("solid", fgColor="7F0000")),
        (0.95, 1.00, PatternFill("solid", fgColor="6A0000")),
    ]

    def _pct_fill(v):
        if v is None or v == "":
            return None
        try:
            x = float(v)
        except Exception:
            return None
        if x <= 0:
            return None
        if x < 0:
            x = 0.0
        if x > 1:
            x = 1.0
        for lo, hi, fill in pct_bins:
            if fill is None:
                continue
            if (lo <= x < hi) or (hi == 1.00 and lo <= x <= hi):
                return fill
        return None

    # GP heatmap
    if gp_idx:
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=gp_idx)
            try:
                v = float(cell.value or 0)
            except Exception:
                continue
            if v <= 0:
                continue
            if v >= 20:
                cell.fill = gp_fill_20p
            elif 16 <= v <= 19:
                cell.fill = gp_fill_16_19
            elif 11 <= v <= 15:
                cell.fill = gp_fill_11_15
            elif 6 <= v <= 10:
                cell.fill = gp_fill_6_10
            elif 1 <= v <= 5:
                cell.fill = gp_fill_1_5

    # % heatmap (GB-/FB- only)
    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(row=2, column=c).value or "").strip()
            if not (h.startswith("GB-") or h.startswith("FB-")):
                continue
            cell = ws.cell(row=r, column=c)
            f = _pct_fill(cell.value)
            if f:
                cell.fill = f

    # Watermark
    try:
        ws.oddHeader.center.text = "RP Spray Analytics"
        ws.oddHeader.center.font = "Tahoma,Bold"
        ws.oddHeader.center.size = 14
        ws.oddHeader.center.color = "808080"
    except Exception:
        pass

    # Print setup
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER

    # -------------------------------------------------
    # ✅ INDIVIDUAL SPRAY CHART TABS (one tab per player)
    # -------------------------------------------------
    used_names = set(sh.title for sh in writer.book.worksheets)

    export_players = display_players[:] if isinstance(display_players, list) else list(season_players.keys())
    export_players = [p for p in export_players if p in season_players]

    for p in export_players:
        tab_name = _safe_sheet_name(p, used_names)
        _build_individual_spray_sheet(
            writer.book,
            tab_name,
            p,
            (season_players.get(p) or {}),
            ""
        )

    # -------------------------------------------------
    # ✅ BLANK INDIVIDUAL TEMPLATE (ALWAYS LAST — EXPORT ONLY)
    # -------------------------------------------------
    template_tab = _safe_sheet_name("NEW PLAYER TEMPLATE", used_names)
    _build_individual_spray_sheet(
        writer.book,
        template_tab,
        "",                 
        {},                 
        "",
        template_mode=True
    )


    # -----------------------------
    # COACH NOTES BOX (EXCEL)
    # -----------------------------
    if notes_box_text:
        top_row = ws.max_row + 6
        left_col = 1
        right_col = ws.max_column
        box_height = 10

        ws.merge_cells(
            start_row=top_row,
            start_column=left_col,
            end_row=top_row + box_height - 1,
            end_column=right_col,
        )

        note_cell = ws.cell(row=top_row, column=left_col)
        note_cell.value = f"COACHES NOTES:\n\n{notes_box_text}"
        note_cell.font = Font(size=12)
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")

        for rr in range(top_row, top_row + box_height):
            ws.row_dimensions[rr].height = 22

        thick = Side(style="thick", color="000000")
        for rr in range(top_row, top_row + box_height):
            for cc in range(left_col, right_col + 1):
                cur = ws.cell(row=rr, column=cc).border
                ws.cell(row=rr, column=cc).border = Border(
                    left=thick if cc == left_col else cur.left,
                    right=thick if cc == right_col else cur.right,
                    top=thick if rr == top_row else cur.top,
                    bottom=thick if rr == top_row + box_height - 1 else cur.bottom,
                )

# ✅ AFTER writer closes: pull bytes
out.seek(0)
excel_bytes = out.getvalue()

# Use the SAME formatted XLSX bytes for Google Sheets
gs_bytes = excel_bytes


with st.container():
    col_dl1, col_dl2, col_dl3 = st.columns([1, 1, 1], gap="small")

with col_dl1:
    st.download_button(
        label="📊 Download Season Report (Excel - Formatted)",
        data=excel_bytes,
        file_name=f"{TEAM_CODE}_{safe_team}_Season_Spray_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_season_excel_{TEAM_CODE}_{_RP_RUN_NONCE}",
        use_container_width=True,
    )

with col_dl2:
    st.download_button(
        label="🟩 Download Season Report (Google Sheets – Formatted)",
        data=gs_bytes,
        file_name=f"{TEAM_CODE}_{safe_team}_Season_Spray_Report_GoogleSheets.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_season_gs_{TEAM_CODE}_{_RP_RUN_NONCE}",
        use_container_width=True,
    )
    st.caption("To open: sheets.google.com → File → Import → Upload.")

with col_dl3:
    st.download_button(
        label="📄 Download Season Report (CSV – Raw Data)",
        data=csv_bytes,
        file_name=f"{TEAM_CODE}_{safe_team}_Season_Spray_Report.csv",
        mime="text/csv",
        key=f"dl_season_csv_{TEAM_CODE}_{_RP_RUN_NONCE}",
        use_container_width=True,
    )


# ✅ IMPORTANT: read bytes AFTER writer closes (after this with-block ends)
out.seek(0)
excel_bytes = out.getvalue()

# Use the SAME formatted XLSX bytes for Google Sheets
gs_bytes = excel_bytes


# -----------------------------
# FOOTER (Copyright)
# -----------------------------
st.markdown(
    """
    <style>
    .rp-footer {
        margin-top: 40px;
        padding-top: 12px;
        border-top: 1px solid rgba(0,0,0,0.12);
        text-align: center;
        font-size: 0.85rem;
        color: rgba(0,0,0,0.55);
    }
    </style>

    <div class="rp-footer">
        © 2026 RP Spray Analytics. All rights reserved.<br>
        Proprietary software. Unauthorized copying, redistribution, or reverse engineering prohibited.
    </div>
    """,
    unsafe_allow_html=True,
)





























































